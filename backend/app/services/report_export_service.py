from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from app.models.monthly_closing import MonthlyClosing
from app.models.print_job import JobStatus, PrintJob

PAGE_WIDTH, PAGE_HEIGHT = A4
PRIMARY = colors.HexColor("#0f6fab")
MUTED = colors.HexColor("#667085")
LIGHT = colors.HexColor("#f2f6f9")
DARK = colors.HexColor("#101828")


def monthly_closing_filename_base(closing: MonthlyClosing) -> str:
    return f"fechamento-{closing.year}-{closing.month:02d}"


def _money(value: float) -> str:
    return f"R$ {float(value or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _truncate(value: str, max_length: int) -> str:
    if len(value) <= max_length:
        return value
    return value[: max_length - 3] + "..."


def _org_name(closing: MonthlyClosing) -> str:
    organization = getattr(closing, "organization", None)
    return getattr(organization, "name", None) or "Sistema de Bilhetagem"


def _job_user_name(job: PrintJob) -> str:
    return job.user.full_name or job.user.username


def _job_department_name(job: PrintJob) -> str:
    return job.user.department.name if job.user and job.user.department else "Sem departamento"


def _job_printer_name(job: PrintJob) -> str:
    return job.printer.name if job.printer else "-"


def _job_status_label(status: JobStatus | str) -> str:
    value = status.value if hasattr(status, "value") else str(status)
    labels = {
        "authorized": "Autorizada",
        "released": "Liberada",
        "pending_release": "Pendente",
        "blocked": "Bloqueada",
        "cancelled": "Cancelada",
    }
    return labels.get(value, value)


def _job_policy_action_label(action: str | None) -> str:
    labels = {
        "allow": "Excecao",
        "block": "Bloqueio",
        "require_release": "Liberacao",
        "force_mono": "Cobrar P&B",
    }
    return labels.get(action or "", "")


def _job_policy_summary(job: PrintJob) -> str:
    if not job.policy_name:
        return ""
    action = _job_policy_action_label(job.policy_action)
    return f"{action}: {job.policy_name}" if action else job.policy_name


def summarize_print_jobs(jobs: list[PrintJob]) -> dict[str, float | int]:
    billable_statuses = {JobStatus.authorized, JobStatus.released}
    saved_statuses = {JobStatus.blocked, JobStatus.cancelled}
    billable_jobs = [job for job in jobs if job.status in billable_statuses]
    saved_jobs = [job for job in jobs if job.status in saved_statuses]
    return {
        "total_jobs": len(jobs),
        "billable_jobs": len(billable_jobs),
        "total_pages": sum(job.pages for job in billable_jobs),
        "mono_pages": sum(job.pages for job in billable_jobs if not job.is_color),
        "color_pages": sum(job.pages for job in billable_jobs if job.is_color),
        "saved_pages": sum(job.pages for job in saved_jobs),
        "total_cost": round(sum(job.cost for job in billable_jobs), 2),
    }


def _draw_header(pdf: canvas.Canvas, closing: MonthlyClosing) -> float:
    pdf.setFillColor(PRIMARY)
    pdf.rect(0, PAGE_HEIGHT - 84, PAGE_WIDTH, 84, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(40, PAGE_HEIGHT - 38, "Fechamento mensal de impressoes")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(40, PAGE_HEIGHT - 57, f"{_org_name(closing)} | Periodo {closing.month:02d}/{closing.year}")
    pdf.drawRightString(PAGE_WIDTH - 40, PAGE_HEIGHT - 57, f"Gerado em {closing.generated_at:%d/%m/%Y %H:%M}")
    return PAGE_HEIGHT - 112


def _draw_metric(pdf: canvas.Canvas, x: float, y: float, w: float, title: str, value: str, subtitle: str = "") -> None:
    pdf.setFillColor(LIGHT)
    pdf.roundRect(x, y - 58, w, 58, 6, fill=1, stroke=0)
    pdf.setFillColor(MUTED)
    pdf.setFont("Helvetica", 8)
    pdf.drawString(x + 10, y - 16, title.upper())
    pdf.setFillColor(DARK)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(x + 10, y - 38, value)
    if subtitle:
        pdf.setFillColor(MUTED)
        pdf.setFont("Helvetica", 7)
        pdf.drawString(x + 10, y - 50, subtitle)


def _new_page(pdf: canvas.Canvas, closing: MonthlyClosing) -> float:
    pdf.showPage()
    return _draw_header(pdf, closing)


def _draw_section(pdf: canvas.Canvas, closing: MonthlyClosing, y: float, title: str, rows: list[dict], limit: int = 8) -> float:
    if y < 160:
        y = _new_page(pdf, closing)
    pdf.setFillColor(DARK)
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(40, y, title)
    y -= 18
    pdf.setFillColor(LIGHT)
    pdf.rect(40, y - 16, PAGE_WIDTH - 80, 18, fill=1, stroke=0)
    pdf.setFillColor(MUTED)
    pdf.setFont("Helvetica-Bold", 8)
    pdf.drawString(48, y - 10, "Nome")
    pdf.drawRightString(330, y - 10, "Paginas")
    pdf.drawRightString(385, y - 10, "P&B")
    pdf.drawRightString(435, y - 10, "Cor")
    pdf.drawRightString(500, y - 10, "Custo/Pag.")
    pdf.drawRightString(PAGE_WIDTH - 48, y - 10, "Custo")
    y -= 24
    pdf.setFont("Helvetica", 8)
    for row in rows[:limit]:
        if y < 72:
            y = _new_page(pdf, closing)
        pdf.setFillColor(DARK)
        name = str(row.get("name", "-"))
        if len(name) > 44:
            name = name[:41] + "..."
        pdf.drawString(48, y, name)
        pdf.drawRightString(330, y, str(row.get("pages", 0)))
        pdf.drawRightString(385, y, str(row.get("mono_pages", 0)))
        pdf.drawRightString(435, y, str(row.get("color_pages", 0)))
        pdf.drawRightString(500, y, _money(float(row.get("cost_per_page", 0))))
        pdf.drawRightString(PAGE_WIDTH - 48, y, _money(float(row.get("cost", 0))))
        y -= 16
    return y - 12


def render_monthly_closing_pdf(closing: MonthlyClosing) -> bytes:
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    pdf.setTitle("Fechamento Mensal")
    y = _draw_header(pdf, closing)

    col_w = (PAGE_WIDTH - 100) / 4
    _draw_metric(pdf, 40, y, col_w, "Paginas cobraveis", str(closing.total_pages), f"{closing.billable_jobs} trabalho(s)")
    _draw_metric(pdf, 50 + col_w, y, col_w, "Custo total", _money(closing.total_cost), "Base para cobranca")
    _draw_metric(pdf, 60 + col_w * 2, y, col_w, "Coloridas", str(closing.color_pages), f"P&B: {closing.mono_pages}")
    _draw_metric(pdf, 70 + col_w * 3, y, col_w, "Paginas salvas", str(closing.blocked_pages), f"{closing.blocked_jobs} bloqueio(s)")
    y -= 92

    eco = closing.snapshot.get("eco", {})
    pdf.setFillColor(colors.HexColor("#e9f8f1"))
    pdf.roundRect(40, y - 44, PAGE_WIDTH - 80, 44, 6, fill=1, stroke=0)
    pdf.setFillColor(colors.HexColor("#067647"))
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(52, y - 15, "Indicadores ambientais estimados")
    pdf.setFont("Helvetica", 9)
    pdf.drawString(
        52,
        y - 32,
        f"CO2 evitado: {eco.get('co2_saved_g', 0)} g  |  Agua preservada: {eco.get('water_saved_l', 0)} L  |  Arvores salvas: {eco.get('trees_saved', 0)}",
    )
    y -= 72

    y = _draw_section(pdf, closing, y, "Ranking por impressora", closing.snapshot.get("by_printer", []))
    y = _draw_section(pdf, closing, y, "Ranking por usuario", closing.snapshot.get("by_user", []))
    y = _draw_section(pdf, closing, y, "Consumo por departamento", closing.snapshot.get("by_department", []))
    y = _draw_section(pdf, closing, y, "Colorido x preto e branco", closing.snapshot.get("by_type", []), limit=4)

    pdf.setFillColor(MUTED)
    pdf.setFont("Helvetica", 7)
    pdf.drawString(40, 36, "Snapshot congelado: valores historicos nao mudam quando usuarios, impressoras ou custos forem editados depois.")
    pdf.save()
    return buffer.getvalue()


def render_print_jobs_pdf(jobs: list[PrintJob], title: str = "Relatorio de impressoes") -> bytes:
    summary = summarize_print_jobs(jobs)
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    pdf.setTitle(title)

    pdf.setFillColor(PRIMARY)
    pdf.rect(0, PAGE_HEIGHT - 84, PAGE_WIDTH, 84, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(40, PAGE_HEIGHT - 38, title)
    pdf.setFont("Helvetica", 10)
    pdf.drawString(40, PAGE_HEIGHT - 57, "Historico filtrado de trabalhos de impressao")
    y = PAGE_HEIGHT - 112

    col_w = (PAGE_WIDTH - 100) / 4
    _draw_metric(pdf, 40, y, col_w, "Trabalhos", str(summary["total_jobs"]), f"{summary['billable_jobs']} cobraveis")
    _draw_metric(pdf, 50 + col_w, y, col_w, "Paginas", str(summary["total_pages"]), f"{summary['saved_pages']} salvas")
    _draw_metric(pdf, 60 + col_w * 2, y, col_w, "Coloridas", str(summary["color_pages"]), f"P&B: {summary['mono_pages']}")
    _draw_metric(pdf, 70 + col_w * 3, y, col_w, "Custo", _money(float(summary["total_cost"])), "Base filtrada")
    y -= 90

    pdf.setFillColor(DARK)
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(40, y, "Trabalhos recentes")
    y -= 18
    pdf.setFillColor(LIGHT)
    pdf.rect(40, y - 16, PAGE_WIDTH - 80, 18, fill=1, stroke=0)
    pdf.setFillColor(MUTED)
    pdf.setFont("Helvetica-Bold", 7)
    pdf.drawString(48, y - 10, "Data")
    pdf.drawString(112, y - 10, "Usuario")
    pdf.drawString(215, y - 10, "Impressora")
    pdf.drawString(335, y - 10, "Documento")
    pdf.drawRightString(475, y - 10, "Pag.")
    pdf.drawString(490, y - 10, "Status/Politica")
    y -= 24
    pdf.setFont("Helvetica", 7)

    for job in jobs[:90]:
        if y < 60:
            pdf.showPage()
            y = PAGE_HEIGHT - 52
            pdf.setFillColor(DARK)
            pdf.setFont("Helvetica-Bold", 10)
            pdf.drawString(40, y, title)
            y -= 24
            pdf.setFont("Helvetica", 7)
        pdf.setFillColor(DARK)
        user_name = _truncate(_job_user_name(job), 24)
        printer_name = _truncate(_job_printer_name(job), 28)
        document_name = _truncate(job.document_name or "-", 28)
        pdf.drawString(48, y, job.submitted_at.strftime("%d/%m/%y %H:%M"))
        pdf.drawString(112, y, user_name)
        pdf.drawString(215, y, printer_name)
        pdf.drawString(335, y, document_name)
        pdf.drawRightString(475, y, str(job.pages))
        status_policy = _job_status_label(job.status)
        policy_summary = _job_policy_summary(job)
        if policy_summary:
            status_policy = f"{status_policy} | {policy_summary}"
        pdf.drawString(490, y, _truncate(status_policy, 24))
        y -= 14

    pdf.setFillColor(MUTED)
    pdf.setFont("Helvetica", 7)
    pdf.drawString(40, 36, "Relatorio operacional filtrado. Fechamentos mensais permanecem como snapshot congelado para cobranca.")
    pdf.save()
    return buffer.getvalue()


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="0F6FAB")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D0D5DD")
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width, 12), 42)


def render_monthly_closing_xlsx(closing: MonthlyClosing) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    summary.append(["Indicador", "Valor"])
    summary.append(["Empresa", _org_name(closing)])
    summary.append(["Periodo", f"{closing.month:02d}/{closing.year}"])
    summary.append(["Trabalhos", closing.total_jobs])
    summary.append(["Trabalhos cobraveis", closing.billable_jobs])
    summary.append(["Trabalhos pendentes", closing.pending_jobs])
    summary.append(["Trabalhos bloqueados", closing.blocked_jobs])
    summary.append(["Paginas cobraveis", closing.total_pages])
    summary.append(["Paginas P&B", closing.mono_pages])
    summary.append(["Paginas coloridas", closing.color_pages])
    summary.append(["Paginas salvas", closing.blocked_pages])
    summary.append(["Custo total", closing.total_cost])
    eco = closing.snapshot.get("eco", {})
    summary.append(["CO2 evitado (g)", eco.get("co2_saved_g", 0)])
    summary.append(["Agua preservada (L)", eco.get("water_saved_l", 0)])
    summary.append(["Arvores salvas", eco.get("trees_saved", 0)])
    summary["B12"].number_format = '"R$" #,##0.00'
    _style_sheet(summary)

    for sheet_name, key in (("Usuarios", "by_user"), ("Departamentos", "by_department"), ("Impressoras", "by_printer"), ("Tipo", "by_type")):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(["Nome", "Trabalhos", "Paginas", "P&B", "Coloridas", "Custo", "Custo/Pag."])
        for row in closing.snapshot.get(key, []):
            sheet.append([row["name"], row["jobs"], row["pages"], row["mono_pages"], row["color_pages"], row["cost"], row.get("cost_per_page", 0)])
        for row in sheet.iter_rows(min_row=2, min_col=6, max_col=7):
            for cell in row:
                cell.number_format = '"R$" #,##0.00'
        _style_sheet(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def render_print_jobs_xlsx(jobs: list[PrintJob]) -> bytes:
    summary = summarize_print_jobs(jobs)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Impressoes"
    sheet.append(["Data", "Usuario", "Departamento", "Impressora", "Documento", "Paginas", "Cor", "Status", "Custo", "Politica", "Acao Politica", "Motivo"])
    for job in jobs:
        sheet.append(
            [
                job.submitted_at.isoformat(),
                _job_user_name(job),
                _job_department_name(job),
                _job_printer_name(job),
                job.document_name or "",
                job.pages,
                "Colorido" if job.is_color else "P&B",
                _job_status_label(job.status),
                job.cost,
                job.policy_name or "",
                _job_policy_action_label(job.policy_action),
                job.reason or "",
            ]
        )
    for row in sheet.iter_rows(min_row=2, min_col=9, max_col=9):
        row[0].number_format = '"R$" #,##0.00'
    _style_sheet(sheet)

    summary_sheet = workbook.create_sheet("Resumo")
    summary_sheet.append(["Indicador", "Valor"])
    summary_sheet.append(["Trabalhos filtrados", summary["total_jobs"]])
    summary_sheet.append(["Trabalhos cobraveis", summary["billable_jobs"]])
    summary_sheet.append(["Paginas cobraveis", summary["total_pages"]])
    summary_sheet.append(["Paginas P&B", summary["mono_pages"]])
    summary_sheet.append(["Paginas coloridas", summary["color_pages"]])
    summary_sheet.append(["Paginas salvas", summary["saved_pages"]])
    summary_sheet.append(["Custo filtrado", summary["total_cost"]])
    summary_sheet["B8"].number_format = '"R$" #,##0.00'
    _style_sheet(summary_sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
