from datetime import datetime, timezone
from io import BytesIO

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.api.routes.printers import delete_printer_endpoint
from app.api.routes.reports import export_monthly_closing, export_report, generate_monthly_closing
from app.api.routes.settings import get_monthly_report_email_settings_endpoint, update_monthly_report_email_settings_endpoint
from app.api.routes.users import delete_user_endpoint
from app.models.department import Department
from app.models.audit_log import AuditLog
from app.models.organization import Organization
from app.models.print_job import JobStatus, PrintJob
from app.models.printer import Printer
from app.schemas.settings import MonthlyReportEmailSettings
from app.schemas.report import MonthlyClosingCreate, MonthlyClosingRead
from app.services.email_service import send_due_monthly_report_email, send_monthly_closing_email
from app.services.email_scheduler import send_due_monthly_reports_once
from app.models.user import User, UserRole
from app.services.monthly_closing_service import create_monthly_closing


def _seed_job_data(db_session: Session) -> tuple[User, Printer]:
    department = Department(organization_id=1, name="Financeiro", cost_center="CC-FIN")
    user = User(username="ana", full_name="Ana Financeiro", role=UserRole.user, department=department, is_active=True, organization_id=1)
    printer = Printer(organization_id=1, name="KONICA_FECHAMENTO", is_color=True, cost_mono=0.05, cost_color=0.25)
    db_session.add_all([department, user, printer])
    db_session.flush()
    jobs = [
        PrintJob(
            organization_id=1,
            user_id=user.id,
            printer_id=printer.id,
            pages=10,
            is_color=False,
            cost=0.50,
            status=JobStatus.authorized,
            reason="Cobrado como P&B pela politica: Cobrar colorido como PB",
            policy_name="Cobrar colorido como PB",
            policy_action="force_mono",
            submitted_at=datetime(2026, 5, 10, 10, tzinfo=timezone.utc),
        ),
        PrintJob(
            organization_id=1,
            user_id=user.id,
            printer_id=printer.id,
            pages=4,
            is_color=True,
            cost=1.00,
            status=JobStatus.released,
            submitted_at=datetime(2026, 5, 11, 10, tzinfo=timezone.utc),
        ),
        PrintJob(
            organization_id=1,
            user_id=user.id,
            printer_id=printer.id,
            pages=3,
            is_color=True,
            cost=0.75,
            status=JobStatus.blocked,
            reason="Colorido bloqueado",
            policy_name="Bloquear colorido",
            policy_action="block",
            submitted_at=datetime(2026, 5, 12, 10, tzinfo=timezone.utc),
        ),
        PrintJob(
            organization_id=1,
            user_id=user.id,
            printer_id=printer.id,
            pages=5,
            is_color=True,
            cost=1.25,
            status=JobStatus.pending_release,
            reason="Liberacao exigida",
            policy_name="Liberacao segura",
            policy_action="require_release",
            submitted_at=datetime(2026, 5, 13, 10, tzinfo=timezone.utc),
        ),
        PrintJob(
            organization_id=1,
            user_id=user.id,
            printer_id=printer.id,
            pages=99,
            is_color=False,
            cost=4.95,
            status=JobStatus.authorized,
            submitted_at=datetime(2026, 6, 1, 10, tzinfo=timezone.utc),
        ),
    ]
    db_session.add_all(jobs)
    db_session.commit()
    return user, printer


def test_monthly_closing_freezes_commercial_snapshot(db_session: Session):
    user, printer = _seed_job_data(db_session)
    organization = db_session.get(Organization, 1)
    organization.name = "Cliente Fechamento"
    organization.slug = "cliente-fechamento"
    organization.billing_plan = "professional"
    organization.billing_status = "active"
    organization.contracted_printer_limit = 2
    db_session.commit()

    closing = create_monthly_closing(db_session, organization_id=1, year=2026, month=5)

    assert closing.total_jobs == 4
    assert closing.billable_jobs == 2
    assert closing.pending_jobs == 1
    assert closing.total_pages == 14
    assert closing.mono_pages == 10
    assert closing.color_pages == 4
    assert closing.blocked_pages == 3
    assert closing.total_cost == 1.5
    assert closing.snapshot["totals"]["released_jobs"] == 1
    assert closing.snapshot["totals"]["pending_pages"] == 5
    assert closing.snapshot["totals"]["pending_cost"] == 1.25
    assert closing.snapshot["totals"]["blocked_cost"] == 0.75
    assert closing.snapshot["totals"]["cost_per_page"] == 0.11
    assert closing.snapshot["totals"]["mono_page_share_percent"] == 71.4
    assert closing.snapshot["totals"]["color_page_share_percent"] == 28.6
    assert closing.snapshot["totals"]["saved_page_share_percent"] == 17.6
    assert closing.snapshot["by_user"][0]["name"] == "Ana Financeiro"
    assert closing.snapshot["by_cost_center"][0]["name"] == "CC-FIN"
    assert closing.snapshot["by_cost_center"][0]["pages"] == 14
    assert closing.snapshot["by_printer"][0]["name"] == "KONICA_FECHAMENTO"
    assert closing.snapshot["by_printer"][0]["cost_per_page"] == 0.11
    assert closing.snapshot["by_printer"][0]["page_share_percent"] == 100.0
    assert closing.snapshot["by_printer"][0]["cost_share_percent"] == 100.0
    assert closing.snapshot["by_policy"] == [
        {
            "name": "Bloquear colorido",
            "action": "block",
            "jobs": 1,
            "billable_jobs": 0,
            "pending_jobs": 0,
            "pending_pages": 0,
            "pending_cost": 0.0,
            "blocked_jobs": 1,
            "blocked_cost": 0.75,
            "pages": 0,
            "mono_pages": 0,
            "color_pages": 0,
            "saved_pages": 3,
            "cost": 0.0,
            "cost_per_page": 0.0,
        },
        {
            "name": "Cobrar colorido como PB",
            "action": "force_mono",
            "jobs": 1,
            "billable_jobs": 1,
            "pending_jobs": 0,
            "pending_pages": 0,
            "pending_cost": 0.0,
            "blocked_jobs": 0,
            "blocked_cost": 0.0,
            "pages": 10,
            "mono_pages": 10,
            "color_pages": 0,
            "saved_pages": 0,
            "cost": 0.5,
            "cost_per_page": 0.05,
        },
        {
            "name": "Liberacao segura",
            "action": "require_release",
            "jobs": 1,
            "billable_jobs": 0,
            "pending_jobs": 1,
            "pending_pages": 5,
            "pending_cost": 1.25,
            "blocked_jobs": 0,
            "blocked_cost": 0.0,
            "pages": 0,
            "mono_pages": 0,
            "color_pages": 0,
            "saved_pages": 0,
            "cost": 0.0,
            "cost_per_page": 0.0,
        },
    ]
    assert closing.snapshot["organization"] == {"id": 1, "name": "Cliente Fechamento", "slug": "cliente-fechamento"}
    assert closing.snapshot["contract"] == {
        "billing_plan": "professional",
        "billing_status": "active",
        "contracted_printer_limit": 2,
        "printers_count": 1,
        "active_printers_count": 1,
        "printer_usage_percent": 50.0,
        "printer_limit_status": "ok",
    }
    validated_closing = MonthlyClosingRead.model_validate(closing)
    assert validated_closing.snapshot.organization.slug == "cliente-fechamento"
    assert validated_closing.snapshot.contract.billing_plan == "professional"
    assert validated_closing.snapshot.contract.printer_usage_percent == 50.0
    assert validated_closing.snapshot.totals.released_jobs == 1
    assert validated_closing.snapshot.totals.pending_cost == 1.25
    assert validated_closing.snapshot.totals.blocked_cost == 0.75
    assert validated_closing.snapshot.totals.cost_per_page == 0.11
    assert validated_closing.snapshot.totals.color_page_share_percent == 28.6
    assert validated_closing.snapshot.totals.saved_page_share_percent == 17.6
    assert validated_closing.snapshot.by_policy[0].name == "Bloquear colorido"
    assert validated_closing.snapshot.by_policy[0].blocked_cost == 0.75
    assert validated_closing.snapshot.by_cost_center[0].name == "CC-FIN"
    assert validated_closing.snapshot.by_printer[0].page_share_percent == 100.0
    assert validated_closing.snapshot.by_printer[0].cost_share_percent == 100.0
    assert validated_closing.snapshot.eco.pages_saved == 3

    user.full_name = "Ana Renomeada"
    printer.name = "KONICA_NOVA"
    organization.name = "Cliente Renomeado"
    organization.slug = "cliente-renomeado"
    organization.billing_plan = "enterprise"
    organization.billing_status = "past_due"
    organization.contracted_printer_limit = 1
    db_session.commit()
    same_closing = create_monthly_closing(db_session, organization_id=1, year=2026, month=5)

    assert same_closing.id == closing.id
    assert same_closing.snapshot["by_user"][0]["name"] == "Ana Financeiro"
    assert same_closing.snapshot["by_printer"][0]["name"] == "KONICA_FECHAMENTO"
    assert same_closing.snapshot["organization"] == {"id": 1, "name": "Cliente Fechamento", "slug": "cliente-fechamento"}
    assert same_closing.snapshot["contract"]["billing_plan"] == "professional"
    assert same_closing.snapshot["contract"]["billing_status"] == "active"
    assert same_closing.snapshot["contract"]["contracted_printer_limit"] == 2


def test_monthly_closing_rejects_invalid_period(db_session: Session):
    with pytest.raises(ValueError, match="Mes do fechamento deve estar entre 1 e 12"):
        create_monthly_closing(db_session, organization_id=1, year=2026, month=13)

    with pytest.raises(ValueError, match="Ano do fechamento deve estar entre 2000 e 2100"):
        create_monthly_closing(db_session, organization_id=1, year=1999, month=5)


def test_user_with_print_history_cannot_be_deleted(db_session: Session):
    user, _ = _seed_job_data(db_session)
    admin = User(username="delete-history-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(admin)
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        delete_user_endpoint(user.id, db=db_session, actor=admin)

    assert exc.value.status_code == 409
    assert db_session.get(User, user.id) is not None
    assert db_session.query(PrintJob).filter(PrintJob.user_id == user.id).count() == 5


def test_printer_with_print_history_cannot_be_deleted(db_session: Session):
    _, printer = _seed_job_data(db_session)
    admin = User(username="delete-printer-history-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(admin)
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        delete_printer_endpoint(printer.id, db=db_session, actor=admin)

    assert exc.value.status_code == 409
    assert db_session.get(Printer, printer.id) is not None
    assert db_session.query(PrintJob).filter(PrintJob.printer_id == printer.id).count() == 5


def test_monthly_closing_export_xlsx(db_session: Session):
    _seed_job_data(db_session)
    organization = db_session.get(Organization, 1)
    organization.name = "Cliente XLSX"
    organization.billing_plan = "enterprise"
    organization.billing_status = "active"
    organization.contracted_printer_limit = 3
    db_session.commit()
    closing = create_monthly_closing(db_session, organization_id=1, year=2026, month=5)
    organization.name = "Cliente XLSX Renomeado"
    actor = User(username="report-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(actor)
    db_session.commit()

    response = export_monthly_closing(closing_id=closing.id, format="xlsx", db=db_session, actor=actor)

    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response.body.startswith(b"PK")
    workbook = load_workbook(BytesIO(response.body), data_only=True)
    assert workbook.sheetnames == ["Resumo", "Usuarios", "Departamentos", "Centros de Custo", "Impressoras", "Tipo", "Politicas"]
    assert workbook["Resumo"]["B2"].value == "Cliente XLSX"
    assert workbook["Resumo"]["A6"].value == "Trabalhos liberados"
    assert workbook["Resumo"]["B6"].value == 1
    assert workbook["Resumo"]["A8"].value == "Paginas pendentes"
    assert workbook["Resumo"]["B8"].value == 5
    assert workbook["Resumo"]["A9"].value == "Custo pendente"
    assert workbook["Resumo"]["B9"].value == 1.25
    assert workbook["Resumo"]["A15"].value == "Custo bloqueado estimado"
    assert workbook["Resumo"]["B15"].value == 0.75
    assert workbook["Resumo"]["A16"].value == "Custo total"
    assert workbook["Resumo"]["B16"].value == 1.5
    assert workbook["Resumo"]["A17"].value == "Custo medio por pagina"
    assert workbook["Resumo"]["B17"].value == 0.11
    assert workbook["Resumo"]["A18"].value == "Paginas P&B (%)"
    assert workbook["Resumo"]["B18"].value == 71.4
    assert workbook["Resumo"]["A19"].value == "Paginas coloridas (%)"
    assert workbook["Resumo"]["B19"].value == 28.6
    assert workbook["Resumo"]["A20"].value == "Paginas salvas (%)"
    assert workbook["Resumo"]["B20"].value == 17.6
    assert workbook["Resumo"]["A24"].value == "Plano"
    assert workbook["Resumo"]["B24"].value == "Enterprise"
    assert workbook["Resumo"]["A25"].value == "Status comercial"
    assert workbook["Resumo"]["B25"].value == "Em dia"
    assert workbook["Resumo"]["A26"].value == "Limite contratado de impressoras"
    assert workbook["Resumo"]["B26"].value == 3
    assert workbook["Resumo"]["A29"].value == "Uso do contrato de impressoras (%)"
    assert workbook["Resumo"]["B29"].value == 33.3
    assert workbook["Impressoras"]["A2"].value == "KONICA_FECHAMENTO"
    assert workbook["Impressoras"]["G1"].value == "Custo/Pag."
    assert workbook["Impressoras"]["G2"].value == 0.11
    assert workbook["Impressoras"]["H1"].value == "% Paginas"
    assert workbook["Impressoras"]["H2"].value == 100
    assert workbook["Impressoras"]["I1"].value == "% Custo"
    assert workbook["Impressoras"]["I2"].value == 100
    assert workbook["Centros de Custo"]["A2"].value == "CC-FIN"
    assert workbook["Centros de Custo"]["C2"].value == 14
    assert workbook["Politicas"]["A1"].value == "Politica"
    assert workbook["Politicas"]["A2"].value == "Bloquear colorido"
    assert workbook["Politicas"]["B2"].value == "Bloqueio"
    assert workbook["Politicas"]["G2"].value == 0
    assert workbook["Politicas"]["H2"].value == 1
    assert workbook["Politicas"]["I2"].value == 0.75
    assert workbook["Politicas"]["M2"].value == 3
    audit = db_session.query(AuditLog).filter(AuditLog.action == "monthly_closing_exported", AuditLog.entity_id == closing.id).one()
    assert audit.log_metadata == {"format": "xlsx"}


def test_monthly_closing_export_pdf(db_session: Session):
    _seed_job_data(db_session)
    closing = create_monthly_closing(db_session, organization_id=1, year=2026, month=5)
    actor = User(username="report-pdf-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(actor)
    db_session.commit()

    response = export_monthly_closing(closing_id=closing.id, format="pdf", db=db_session, actor=actor)

    assert response.media_type == "application/pdf"
    assert response.body.startswith(b"%PDF")


def test_report_export_applies_department_filter(db_session: Session):
    user, printer = _seed_job_data(db_session)
    other_department = Department(organization_id=1, name="Juridico")
    other_user = User(username="bia", full_name="Bia Juridico", role=UserRole.user, department=other_department, is_active=True, organization_id=1)
    other_printer = Printer(organization_id=1, name="HP_JURIDICO", is_color=False, cost_mono=0.05, cost_color=0.25)
    db_session.add_all([other_department, other_user, other_printer])
    db_session.flush()
    db_session.add(
        PrintJob(
            organization_id=1,
            user_id=other_user.id,
            printer_id=other_printer.id,
            pages=2,
            is_color=False,
            cost=0.10,
            status=JobStatus.authorized,
            submitted_at=datetime(2026, 5, 13, 10, tzinfo=timezone.utc),
        )
    )
    actor = User(username="report-filter-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(actor)
    db_session.commit()

    response = export_report(format="xlsx", department_id=user.department_id, db=db_session, actor=actor)

    workbook = load_workbook(BytesIO(response.body), data_only=True)
    assert workbook.sheetnames == ["Impressoes", "Resumo"]
    sheet = workbook.active
    exported_users = [row[1].value for row in sheet.iter_rows(min_row=2)]
    assert set(exported_users) == {"Ana Financeiro"}
    assert len(exported_users) == 5
    assert sheet["C2"].value == "Financeiro"
    assert sheet["D2"].value == "CC-FIN"
    assert sheet["J2"].value in {0.5, 1.0, 0.75, 4.95}
    policy_rows = [row for row in sheet.iter_rows(min_row=2, values_only=True) if row[10] == "Cobrar colorido como PB"]
    assert len(policy_rows) == 1
    assert policy_rows[0][11] == "Cobrar P&B"
    assert policy_rows[0][12] == "Cobrado como P&B pela politica: Cobrar colorido como PB"
    assert workbook["Resumo"]["A8"].value == "Custo filtrado"
    assert workbook["Resumo"]["B8"].value == 6.45
    assert workbook["Resumo"]["A10"].value == "Filtros aplicados"
    assert workbook["Resumo"]["A11"].value == "Departamento"
    assert workbook["Resumo"]["B11"].value == "Financeiro"

    audit = db_session.query(AuditLog).filter(AuditLog.action == "report_exported").one()
    assert audit.actor_user_id == actor.id
    assert audit.log_metadata["format"] == "xlsx"
    assert audit.log_metadata["rows"] == 5
    assert audit.log_metadata["filters"]["department_id"] == user.department_id
    assert audit.log_metadata["filter_summary"] == {"Departamento": "Financeiro"}


def test_report_export_applies_cost_center_filter(db_session: Session):
    _seed_job_data(db_session)
    other_department = Department(organization_id=1, name="Juridico", cost_center="CC-JUR")
    other_user = User(username="bia-cc", full_name="Bia Juridico", role=UserRole.user, department=other_department, is_active=True, organization_id=1)
    other_printer = Printer(organization_id=1, name="HP_JURIDICO_CC", is_color=False, cost_mono=0.05, cost_color=0.25)
    db_session.add_all([other_department, other_user, other_printer])
    db_session.flush()
    db_session.add(
        PrintJob(
            organization_id=1,
            user_id=other_user.id,
            printer_id=other_printer.id,
            pages=2,
            is_color=False,
            cost=0.10,
            status=JobStatus.authorized,
            submitted_at=datetime(2026, 5, 13, 10, tzinfo=timezone.utc),
        )
    )
    actor = User(username="report-cost-center-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(actor)
    db_session.commit()

    response = export_report(format="xlsx", cost_center="CC-FIN", db=db_session, actor=actor)

    workbook = load_workbook(BytesIO(response.body), data_only=True)
    sheet = workbook["Impressoes"]
    exported_users = [row[1].value for row in sheet.iter_rows(min_row=2)]
    assert set(exported_users) == {"Ana Financeiro"}
    assert len(exported_users) == 5
    assert {row[3].value for row in sheet.iter_rows(min_row=2)} == {"CC-FIN"}
    assert workbook["Resumo"]["A11"].value == "Centro de Custo"
    assert workbook["Resumo"]["B11"].value == "CC-FIN"

    audit = db_session.query(AuditLog).filter(AuditLog.action == "report_exported").one()
    assert audit.log_metadata["filters"]["cost_center"] == "CC-FIN"
    assert audit.log_metadata["filter_summary"] == {"Centro de Custo": "CC-FIN"}


def test_report_export_audit_and_summary_warn_when_limited(db_session: Session):
    user = User(username="bulk-report-user", full_name="Usuario Relatorio Grande", role=UserRole.user, is_active=True, organization_id=1)
    printer = Printer(organization_id=1, name="KONICA_RELATORIO_GRANDE", is_color=False, cost_mono=0.05, cost_color=0.25)
    actor = User(username="bulk-report-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add_all([user, printer, actor])
    db_session.flush()
    db_session.bulk_save_objects(
        [
            PrintJob(
                organization_id=1,
                user_id=user.id,
                printer_id=printer.id,
                document_name=f"Relatorio {index}",
                pages=1,
                is_color=False,
                cost=0.05,
                status=JobStatus.authorized,
                submitted_at=datetime(2026, 5, 13, 10, index % 60, tzinfo=timezone.utc),
            )
            for index in range(5001)
        ]
    )
    db_session.commit()

    response = export_report(format="xlsx", db=db_session, actor=actor)

    workbook = load_workbook(BytesIO(response.body), data_only=True)
    sheet = workbook["Impressoes"]
    summary = workbook["Resumo"]
    exported_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(exported_rows) == 5000
    assert summary["A11"].value == "Limite da Exportacao"
    assert summary["B11"].value == "5000 de 5001 registros exportados"
    audit = db_session.query(AuditLog).filter(AuditLog.action == "report_exported").one()
    assert audit.log_metadata["rows"] == 5000
    assert audit.log_metadata["total_matching_rows"] == 5001
    assert audit.log_metadata["limit"] == 5000
    assert audit.log_metadata["truncated"] is True
    assert audit.log_metadata["filter_summary"] == {"Limite da Exportacao": "5000 de 5001 registros exportados"}


def test_report_export_xlsx_escapes_formula_like_text(db_session: Session):
    department = Department(organization_id=1, name="+Financeiro", cost_center="-CC")
    user = User(username="formula-user", full_name="=Ana Financeiro", role=UserRole.user, department=department, is_active=True, organization_id=1)
    printer = Printer(organization_id=1, name="@KONICA_FORMULA", is_color=True, cost_mono=0.05, cost_color=0.25)
    actor = User(username="report-formula-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add_all([department, user, printer, actor])
    db_session.flush()
    db_session.add(
        PrintJob(
            organization_id=1,
            user_id=user.id,
            printer_id=printer.id,
            document_name="-Relatorio.xlsx",
            pages=1,
            is_color=False,
            cost=0.05,
            status=JobStatus.authorized,
            policy_name="+Politica",
            policy_action="block",
            reason="@Motivo",
            submitted_at=datetime(2026, 5, 13, 10, tzinfo=timezone.utc),
        )
    )
    db_session.commit()

    response = export_report(format="xlsx", db=db_session, actor=actor)

    workbook = load_workbook(BytesIO(response.body), data_only=False)
    sheet = workbook["Impressoes"]
    assert sheet["B2"].value == "'=Ana Financeiro"
    assert sheet["B2"].data_type == "s"
    assert sheet["C2"].value == "'+Financeiro"
    assert sheet["D2"].value == "'-CC"
    assert sheet["E2"].value == "'@KONICA_FORMULA"
    assert sheet["F2"].value == "'-Relatorio.xlsx"
    assert sheet["K2"].value == "'+Politica"
    assert sheet["M2"].value == "'@Motivo"


def test_report_export_rejects_filters_from_other_organization(db_session: Session):
    other_org = Organization(name="Cliente Relatorio B", slug="cliente-relatorio-b", is_active=True)
    actor = User(username="report-scope-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    other_department = Department(organization=other_org, name="Financeiro", cost_center="CC-OUTRO")
    other_user = User(username="report-scope-user", full_name="Outro Usuario", role=UserRole.user, is_active=True, organization=other_org, department=other_department)
    other_printer = Printer(organization=other_org, name="KONICA OUTRA EMPRESA", is_color=True)
    db_session.add_all([other_org, actor, other_department, other_user, other_printer])
    db_session.commit()

    with pytest.raises(HTTPException) as user_exc:
        export_report(format="xlsx", user_id=other_user.id, db=db_session, actor=actor)
    assert user_exc.value.status_code == 404

    with pytest.raises(HTTPException) as department_exc:
        export_report(format="xlsx", department_id=other_department.id, db=db_session, actor=actor)
    assert department_exc.value.status_code == 404

    with pytest.raises(HTTPException) as cost_center_exc:
        export_report(format="xlsx", cost_center="CC-OUTRO", db=db_session, actor=actor)
    assert cost_center_exc.value.status_code == 404

    with pytest.raises(HTTPException) as printer_exc:
        export_report(format="xlsx", printer_id=other_printer.id, db=db_session, actor=actor)
    assert printer_exc.value.status_code == 404

    assert db_session.query(AuditLog).filter(AuditLog.action == "report_exported").count() == 0


def test_report_export_rejects_zero_filter_ids_without_audit(db_session: Session):
    actor = User(username="report-zero-filter-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(actor)
    db_session.commit()

    with pytest.raises(HTTPException) as user_exc:
        export_report(format="xlsx", user_id=0, db=db_session, actor=actor)
    assert user_exc.value.status_code == 404

    with pytest.raises(HTTPException) as department_exc:
        export_report(format="xlsx", department_id=0, db=db_session, actor=actor)
    assert department_exc.value.status_code == 404

    with pytest.raises(HTTPException) as printer_exc:
        export_report(format="xlsx", printer_id=0, db=db_session, actor=actor)
    assert printer_exc.value.status_code == 404

    assert db_session.query(AuditLog).filter(AuditLog.action == "report_exported").count() == 0


def test_report_export_rejects_invalid_date_range_without_audit(db_session: Session):
    actor = User(username="report-date-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(actor)
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        export_report(
            format="xlsx",
            date_from=datetime(2026, 6, 11, tzinfo=timezone.utc),
            date_to=datetime(2026, 6, 10, tzinfo=timezone.utc),
            db=db_session,
            actor=actor,
        )

    assert exc.value.status_code == 400
    assert db_session.query(AuditLog).filter(AuditLog.action == "report_exported").count() == 0


def test_report_export_pdf_uses_commercial_renderer_and_audit(db_session: Session):
    _seed_job_data(db_session)
    actor = User(username="report-general-pdf-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(actor)
    db_session.commit()

    response = export_report(format="pdf", db=db_session, actor=actor)

    assert response.media_type == "application/pdf"
    assert response.body.startswith(b"%PDF")
    audit = db_session.query(AuditLog).filter(AuditLog.action == "report_exported").one()
    assert audit.actor_user_id == actor.id
    assert audit.log_metadata["format"] == "pdf"
    assert audit.log_metadata["rows"] == 5


def test_monthly_report_email_settings_api(db_session: Session):
    actor = User(username="email-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(actor)
    db_session.commit()

    initial = get_monthly_report_email_settings_endpoint(db=db_session, actor=actor)
    assert initial.enabled is False
    assert initial.day_of_month == 1

    updated = update_monthly_report_email_settings_endpoint(
        payload=MonthlyReportEmailSettings(
            enabled=True,
            recipients="financeiro@example.com; gestao@example.com",
            day_of_month=5,
            include_pdf=True,
            include_xlsx=False,
        ),
        db=db_session,
        actor=actor,
    )

    assert updated.enabled is True
    assert updated.recipients == "financeiro@example.com; gestao@example.com"
    assert updated.day_of_month == 5
    assert updated.include_xlsx is False


def test_generate_monthly_closing_endpoint_writes_audit(db_session: Session):
    _seed_job_data(db_session)
    actor = User(username="closing-audit-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(actor)
    db_session.commit()

    closing = generate_monthly_closing(MonthlyClosingCreate(year=2026, month=5), db=db_session, actor=actor)

    audit = db_session.query(AuditLog).filter(AuditLog.action == "monthly_closing_generated", AuditLog.entity_id == closing.id).one()
    assert audit.actor_user_id == actor.id
    assert audit.log_metadata["year"] == 2026
    assert audit.log_metadata["month"] == 5
    assert audit.log_metadata["total_pages"] == 14
    assert audit.log_metadata["total_cost"] == 1.5


def test_send_monthly_closing_email_with_attachments(db_session: Session, monkeypatch):
    _seed_job_data(db_session)
    closing = create_monthly_closing(db_session, organization_id=1, year=2026, month=5)
    sent_messages = []

    class DummySMTP:
        def __init__(self, host, port, timeout):
            self.host = host
            self.port = port
            self.timeout = timeout

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def starttls(self):
            return None

        def login(self, username, password):
            return None

        def send_message(self, message):
            sent_messages.append(message)

    monkeypatch.setattr("app.services.email_service.settings.smtp_host", "smtp.example.com")
    monkeypatch.setattr("app.services.email_service.smtplib.SMTP", DummySMTP)

    result = send_monthly_closing_email(
        db_session,
        closing,
        recipients="financeiro@example.com,gestao@example.com",
    )

    assert result["sent"] is True
    assert result["recipients"] == ["financeiro@example.com", "gestao@example.com"]
    assert result["attachments"] == ["fechamento-2026-05.pdf", "fechamento-2026-05.xlsx"]
    assert len(sent_messages) == 1
    assert sent_messages[0]["To"] == "financeiro@example.com, gestao@example.com"
    assert [part.get_filename() for part in sent_messages[0].iter_attachments()] == result["attachments"]


def test_due_monthly_report_email_sends_previous_month_once(db_session: Session, monkeypatch):
    _seed_job_data(db_session)
    sent_messages = []

    class DummySMTP:
        def __init__(self, host, port, timeout):
            pass

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def starttls(self):
            return None

        def send_message(self, message):
            sent_messages.append(message)

    monkeypatch.setattr("app.services.email_service.settings.smtp_host", "smtp.example.com")
    monkeypatch.setattr("app.services.email_service.smtplib.SMTP", DummySMTP)
    update_monthly_report_email_settings_endpoint(
        payload=MonthlyReportEmailSettings(
            enabled=True,
            recipients="financeiro@example.com",
            day_of_month=1,
            include_pdf=True,
            include_xlsx=False,
        ),
        db=db_session,
        actor=User(username="email-due-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1),
    )

    first = send_due_monthly_report_email(db_session, organization_id=1, now=datetime(2026, 6, 10, tzinfo=timezone.utc))
    second = send_due_monthly_report_email(db_session, organization_id=1, now=datetime(2026, 6, 10, tzinfo=timezone.utc))

    assert first["sent"] is True
    assert first["period"] == "2026-05"
    assert first["attachments"] == ["fechamento-2026-05.pdf"]
    assert second["sent"] is False
    assert second["reason"] == "Fechamento mensal ja enviado"
    assert len(sent_messages) == 1


def test_monthly_report_email_scheduler_processes_active_organizations(db_session: Session, monkeypatch):
    active_org = Organization(name="Cliente Scheduler", slug="cliente-scheduler", is_active=True)
    inactive_org = Organization(name="Cliente Scheduler Inativo", slug="cliente-scheduler-inativo", is_active=False)
    suspended_org = Organization(name="Cliente Scheduler Suspenso", slug="cliente-scheduler-suspenso", is_active=True, billing_status="suspended")
    past_due_org = Organization(name="Cliente Scheduler Atrasado", slug="cliente-scheduler-atrasado", is_active=True, billing_status="past_due")
    db_session.add_all([active_org, inactive_org, suspended_org, past_due_org])
    db_session.commit()

    called_organization_ids = []

    def fake_send_due(db, organization_id, now=None):
        called_organization_ids.append(organization_id)
        if organization_id == 1:
            return {
                "sent": True,
                "period": "2026-05",
                "closing_id": 99,
                "recipients": ["financeiro@example.com"],
                "attachments": ["fechamento-2026-05.pdf"],
                "reason": None,
            }
        return {"sent": False, "reason": "Envio mensal desativado"}

    monkeypatch.setattr("app.services.email_scheduler.send_due_monthly_report_email", fake_send_due)

    results = send_due_monthly_reports_once(db_session, now=datetime(2026, 6, 10, tzinfo=timezone.utc))

    assert called_organization_ids == [1, active_org.id, past_due_org.id]
    assert [result["organization_slug"] for result in results] == ["default", "cliente-scheduler", "cliente-scheduler-atrasado"]
    audit = db_session.query(AuditLog).filter(AuditLog.action == "monthly_closing_due_email_sent").one()
    assert audit.organization_id == 1
    assert audit.entity_id == 99
    assert audit.log_metadata["automatic"] is True


def test_monthly_report_email_scheduler_audits_failures_per_organization(db_session: Session, monkeypatch):
    failing_org = Organization(name="Cliente Scheduler Falha", slug="cliente-scheduler-falha", is_active=True)
    healthy_org = Organization(name="Cliente Scheduler OK", slug="cliente-scheduler-ok", is_active=True)
    db_session.add_all([failing_org, healthy_org])
    db_session.commit()

    def fake_send_due(db, organization_id, now=None):
        if organization_id == failing_org.id:
            raise RuntimeError("SMTP indisponivel")
        return {"sent": False, "reason": "Envio mensal desativado"}

    monkeypatch.setattr("app.services.email_scheduler.send_due_monthly_report_email", fake_send_due)

    results = send_due_monthly_reports_once(db_session, now=datetime(2026, 6, 10, tzinfo=timezone.utc))

    failure = next(result for result in results if result["organization_id"] == failing_org.id)
    assert failure["sent"] is False
    assert failure["error"] is True
    assert failure["reason"] == "SMTP indisponivel"
    assert {result["organization_id"] for result in results} >= {1, failing_org.id, healthy_org.id}
    audit = db_session.query(AuditLog).filter(AuditLog.action == "monthly_closing_due_email_failed").one()
    assert audit.organization_id == failing_org.id
    assert audit.actor_user_id is None
    assert audit.log_metadata == {"automatic": True, "reason": "SMTP indisponivel"}
