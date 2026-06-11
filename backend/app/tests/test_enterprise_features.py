from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.department import Department
from app.models.agent_queue_action import AgentQueueAction, AgentQueueActionStatus, AgentQueueActionType
from app.models.agent_log import AgentLog
from app.models.audit_log import AuditLog
from app.models.printer import Printer
from app.models.print_agent import PrintAgent
from app.models.printer_alias import PrinterAlias
from app.models.organization import Organization
from app.models.user import User, UserRole
from app.models.print_job import JobStatus, PrintJob
from app.models.print_policy import PolicyAction, PolicyRuleType, PrintPolicy
from app.models.quota import Quota
from app.api.routes.auth import current_auth_context, login
from app.api.routes.agent_updates import list_agents
from app.api.routes.audit_logs import export_audit_logs, list_audit_log_facets, list_audit_logs
from app.api.routes.organizations import create_organization, list_organizations, update_organization
from app.api.routes.reports import export_report
from app.api.routes.printers import bind_printer_alias_endpoint, delete_printer_endpoint, merge_printer_endpoint, update_printer_endpoint, update_printer_status_endpoint
from app.api.routes.settings import sync_ldap_endpoint
from app.schemas.printer import PrinterAliasBind, PrinterStatusUpdate, PrinterUpdate
from app.api.routes.jobs import list_jobs
from app.api.routes.quotas import update_quota
from app.api.routes.departments import create_department, delete_department, list_departments, update_department
from app.api.routes.printers import list_printers
from app.api.routes.users import create_user_endpoint, delete_user_endpoint, list_users, update_user_endpoint
from app.core.security import create_access_token, hash_password, verify_password
from app.core.config import settings
from app.core.deps import get_current_user, require_roles
from app.schemas.department import DepartmentCreate, DepartmentUpdate
from app.schemas.auth import LoginRequest
from app.schemas.quota import QuotaUpdate
from app.schemas.user import UserCreate, UserUpdate
from app.services.report_service import dashboard_metrics
from app.services.monthly_closing_service import build_monthly_snapshot
from app.services.audit_service import write_audit
from app.services.snmp_service import SnmpPrinterStatus, poll_printers_once
from app.services.ldap_service import LDAPUserRecord, sync_ldap_users, test_ldap_connection as check_ldap_connection
from app.seed import ensure_user, validate_seed_password
from app.api.routes.jobs import get_pdf_page_count
from app.services.print_job_service import register_print_job
from app.schemas.job import PrintJobCreate
from app.schemas.organization import OrganizationCreate, OrganizationUpdate
from app.schemas.report import DashboardMetrics
from app.schemas.settings import LDAPSettings


def test_seed_rejects_default_or_placeholder_passwords():
    unsafe_passwords = (
        "",
        "admin",
        "agent",
        "admin12345",
        "agent12345",
        "change-me-admin-password",
        "change-me-agent-password",
        "password",
        "senha123",
        "12345678",
    )
    for password in unsafe_passwords:
        with pytest.raises(RuntimeError):
            validate_seed_password(password, "INITIAL_ADMIN_PASSWORD")

    validate_seed_password("StrongSeedPassword2026", "INITIAL_ADMIN_PASSWORD")


def test_seed_does_not_require_initial_password_for_existing_user(db_session: Session):
    db_session.add(
        User(
            organization_id=1,
            username="admin",
            full_name="Administrador",
            password_hash=hash_password("ExistingAdminPassword2026"),
            role=UserRole.admin,
            is_active=True,
        )
    )
    db_session.commit()

    ensure_user(
        db_session,
        username="admin",
        password="admin12345",
        role=UserRole.admin,
        full_name="Administrador",
    )

    assert db_session.query(User).filter(User.username == "admin").count() == 1


def test_agent_role_can_use_agent_endpoints_but_not_admin_dependencies():
    agent_user = User(username="agent-tech", full_name="Agent Tecnico", role=UserRole.agent, is_active=True, organization_id=1)

    assert require_roles(UserRole.agent, UserRole.admin)(agent_user) == agent_user
    with pytest.raises(HTTPException) as exc:
        require_roles(UserRole.admin)(agent_user)
    assert exc.value.status_code == 403
    with pytest.raises(HTTPException) as human_exc:
        require_roles(UserRole.admin, UserRole.manager, UserRole.user)(agent_user)
    assert human_exc.value.status_code == 403


def test_authenticated_routes_reject_tokens_without_organization_context(db_session: Session):
    user = User(username="token-scope", full_name="Token Scope", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(user)
    db_session.commit()

    valid_token = create_access_token(user.username, {"role": user.role.value, "organization_id": user.organization_id})
    legacy_token = create_access_token(user.username, {"role": user.role.value})
    invalid_org_token = create_access_token(user.username, {"role": user.role.value, "organization_id": "abc"})

    assert get_current_user(valid_token, db_session).id == user.id
    with pytest.raises(HTTPException) as legacy_exc:
        get_current_user(legacy_token, db_session)
    assert legacy_exc.value.status_code == 401

    with pytest.raises(HTTPException) as invalid_exc:
        get_current_user(invalid_org_token, db_session)
    assert invalid_exc.value.status_code == 401


def test_snmp_poll_uses_real_status_payload(db_session: Session, monkeypatch):
    # Mock SessionLocal in snmp_service to return db_session
    import app.services.snmp_service as snmp_mod
    monkeypatch.setattr(snmp_mod, "SessionLocal", lambda: db_session)
    # Prevent the service from closing the test DB session
    monkeypatch.setattr(db_session, "close", lambda: None)
    
    monkeypatch.setattr(
        snmp_mod,
        "fetch_printer_snmp",
        lambda _: SnmpPrinterStatus(
            serial_number="BR123456",
            page_counter=12345,
            toner_levels={"black": 81, "cyan": 72, "magenta": 64, "yellow": 57},
            paper_status="Pronta",
        ),
    )

    printer = Printer(name="HP Lab", location="Lab", ip_address="192.168.1.99")
    db_session.add(printer)
    db_session.commit()
    
    poll_printers_once()
    printer = db_session.query(Printer).filter(Printer.id == printer.id).one()
    
    assert printer.serial_number == "BR123456"
    assert printer.page_counter == 12345
    assert printer.toner_levels == {"black": 81, "cyan": 72, "magenta": 64, "yellow": 57}
    assert printer.toner_level == 57
    assert printer.paper_status == "Pronta"


def test_ldap_user_and_department_sync(db_session: Session, monkeypatch):
    technical_agent = User(
        organization_id=1,
        username="agent",
        full_name="Agente Windows",
        role=UserRole.agent,
        is_active=True,
    )
    db_session.add(technical_agent)
    db_session.commit()
    monkeypatch.setattr(
        "app.services.ldap_service._fetch_ldap_users",
        lambda server, bind_dn, bind_password, search_base: [
            LDAPUserRecord(username="ana.silva", full_name="Ana Silva", department="TI"),
            LDAPUserRecord(username="pedro.santos", full_name="Pedro Santos", department="Financeiro"),
            LDAPUserRecord(username="carla.souza", full_name="Carla Souza", department="Recursos Humanos"),
            LDAPUserRecord(username="marcos.oliveira", full_name="Marcos Oliveira", department="Vendas"),
            LDAPUserRecord(username="agent", full_name="Conta LDAP Agent", department="TI"),
        ],
    )

    result = sync_ldap_users(
        db=db_session,
        server="ldap://localhost:389",
        bind_dn="cn=admin,dc=example,dc=com",
        bind_password="secret",
        search_base="dc=example,dc=com"
    )
    
    assert result["success"] is True
    assert result["total_synced"] == 4
    assert result["new_users"] == 4
    assert result["skipped_users"] == 1
    
    # Verify users and departments are in the DB
    users = db_session.query(User).all()
    usernames = {u.username for u in users}
    assert "ana.silva" in usernames
    assert "pedro.santos" in usernames
    synced_user = db_session.query(User).filter(User.username == "ana.silva").one()
    assert synced_user.password_hash is None
    db_session.refresh(technical_agent)
    assert technical_agent.role == UserRole.agent
    assert technical_agent.full_name == "Agente Windows"
    
    # Verify quotas were initialized
    quota = db_session.query(Quota).filter(Quota.user_id == synced_user.id).first()
    assert quota is not None
    assert quota.monthly_balance == 50.0

    with pytest.raises(ValueError):
        check_ldap_connection("", "", "")


def test_ldap_connection_uses_real_bind_and_unbind(monkeypatch):
    calls = []

    class FakeServer:
        def __init__(self, server, get_info=None, connect_timeout=None):
            calls.append(("server", server, get_info, connect_timeout))

    class FakeConnection:
        def __init__(self, server, user=None, password=None, auto_bind=False, receive_timeout=None):
            calls.append(("connection", user, password, auto_bind, receive_timeout))
            self.bound = True

        def unbind(self):
            calls.append(("unbind",))

    monkeypatch.setattr("app.services.ldap_service.Server", FakeServer)
    monkeypatch.setattr("app.services.ldap_service.Connection", FakeConnection)
    monkeypatch.setattr("app.services.ldap_service.LDAPException", Exception)
    monkeypatch.setattr("app.services.ldap_service.ALL", "ALL")

    assert check_ldap_connection("ldap://ad.local:389", "cn=admin,dc=empresa,dc=local", "secret") is True
    assert calls == [
        ("server", "ldap://ad.local:389", "ALL", 5),
        ("connection", "cn=admin,dc=empresa,dc=local", "secret", True, 15),
        ("unbind",),
    ]


def test_ldap_sync_endpoint_persists_audit_after_service_commit(db_session: Session, monkeypatch):
    actor = User(username="ldap-audit-admin", full_name="LDAP Audit Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(actor)
    db_session.commit()
    monkeypatch.setattr(
        "app.services.ldap_service._fetch_ldap_users",
        lambda server, bind_dn, bind_password, search_base: [
            LDAPUserRecord(username="ana.silva", full_name="Ana Silva", department="TI"),
            LDAPUserRecord(username="pedro.santos", full_name="Pedro Santos", department="Financeiro"),
            LDAPUserRecord(username="carla.souza", full_name="Carla Souza", department="Recursos Humanos"),
            LDAPUserRecord(username="marcos.oliveira", full_name="Marcos Oliveira", department="Vendas"),
        ],
    )
    original_commit = db_session.commit
    commit_count = 0

    def tracked_commit():
        nonlocal commit_count
        commit_count += 1
        original_commit()

    monkeypatch.setattr(db_session, "commit", tracked_commit)

    result = sync_ldap_endpoint(
        payload=LDAPSettings(
            server="ldap://localhost:389",
            bind_dn="cn=admin,dc=example,dc=com",
            bind_password="secret",
            search_base="dc=example,dc=com",
        ),
        db=db_session,
        actor=actor,
    )

    audit = db_session.query(AuditLog).filter(AuditLog.action == "ldap_sync_performed").one()
    assert result["total_synced"] == 4
    assert commit_count == 2
    assert audit.organization_id == actor.organization_id
    assert audit.actor_user_id == actor.id
    assert audit.log_metadata["server"] == "ldap://localhost:389"
    assert audit.log_metadata["new_users"] == 4
    assert audit.log_metadata["total_synced"] == 4


def test_pdf_page_counting_logic():
    # 1. Simple simulated PDF headers with page type definitions
    pdf_content_1 = b"%PDF-1.4\n1 0 obj\n<< /Type /Page >>\nendobj\n2 0 obj\n<< /Type /Page >>\nendobj"
    assert get_pdf_page_count(pdf_content_1) == 2
    
    # 2. PDF headers with /Count
    pdf_content_2 = b"%PDF-1.4\n<< /Type /Pages /Count 5 >>"
    assert get_pdf_page_count(pdf_content_2) == 5
    
    # 3. Fallback default
    assert get_pdf_page_count(b"invalid-bytes") == 1


def test_jobs_with_different_queue_names_share_same_physical_printer(db_session: Session):
    payload_one = PrintJobCreate(
        username="diego",
        printer_name="KONICA MINOLTA C368SeriesPS",
        pages=1,
        is_color=False,
        external_job_id="eventlog:alias-1",
        agent_uid="agent-pc-a",
        computer_name="PC-A",
        queue_name="KONICA MINOLTA C368SeriesPS",
        printer_ip_address="192.168.1.125",
        printer_port_name="IP_192.168.1.125",
        printer_driver_name="KONICA Driver",
        printer_connection_type="network",
        printer_fingerprint="ip:192.168.1.125",
    )
    payload_two = PrintJobCreate(
        username="maria",
        printer_name="Konica Financeiro",
        pages=1,
        is_color=False,
        external_job_id="eventlog:alias-2",
        agent_uid="agent-pc-b",
        computer_name="PC-B",
        queue_name="Konica Financeiro",
        printer_ip_address="192.168.1.125",
        printer_port_name="IP_192.168.1.125",
        printer_driver_name="KONICA Driver",
        printer_connection_type="network",
        printer_fingerprint="ip:192.168.1.125",
    )

    register_print_job(db_session, payload_one)
    register_print_job(db_session, payload_two)

    assert db_session.query(Printer).count() == 1
    printer = db_session.query(Printer).one()
    assert printer.name == "KONICA MINOLTA C368SeriesPS"
    assert len(printer.aliases) == 2


def test_job_updates_printer_ip_when_serial_matches_existing_device(db_session: Session):
    printer = Printer(
        organization_id=1,
        name="KONICA SERIAL IP",
        is_color=True,
        serial_number="SN-IP-001",
        ip_address="192.168.1.10",
    )
    db_session.add(printer)
    db_session.commit()

    register_print_job(
        db_session,
        PrintJobCreate(
            username="diego",
            printer_name="KONICA NOVO IP",
            pages=1,
            is_color=False,
            external_job_id="eventlog:serial-ip-change",
            agent_uid="agent-serial-ip",
            computer_name="PC-SERIAL-IP",
            queue_name="KONICA NOVO IP",
            printer_serial="sn-ip-001",
            printer_ip_address="192.168.1.125",
            printer_connection_type="network",
            printer_fingerprint="serial:sn-ip-001",
        ),
    )

    updated = db_session.get(Printer, printer.id)
    job = db_session.query(PrintJob).one()
    assert db_session.query(Printer).count() == 1
    assert updated.ip_address == "192.168.1.125"
    assert job.printer_id == printer.id


def test_job_reuses_agent_alias_by_normalized_queue_name(db_session: Session):
    printer = Printer(organization_id=1, name="KONICA FINANCEIRO", is_color=True)
    agent = PrintAgent(organization_id=1, agent_uid="agent-normalized-queue", computer_name="PC-FIN")
    db_session.add_all([printer, agent])
    db_session.flush()
    alias = PrinterAlias(
        organization_id=1,
        printer=printer,
        agent=agent,
        queue_name="KONICA Financeiro",
        normalized_queue_name="konica financeiro",
    )
    db_session.add(alias)
    db_session.commit()

    register_print_job(
        db_session,
        PrintJobCreate(
            username="diego",
            printer_name="KONICA FINANCEIRO LOCAL",
            queue_name="  konica   financeiro ",
            pages=1,
            is_color=False,
            agent_uid=agent.agent_uid,
            external_job_id="eventlog:normalized-queue",
        ),
    )

    job = db_session.query(PrintJob).one()
    assert job.printer_id == printer.id
    assert job.printer_alias_id == alias.id
    assert db_session.query(PrinterAlias).count() == 1
    assert db_session.query(Printer).count() == 1


def test_job_does_not_clear_known_alias_metadata_when_payload_is_incomplete(db_session: Session):
    printer = Printer(organization_id=1, name="KONICA METADATA", is_color=True)
    agent = PrintAgent(organization_id=1, agent_uid="agent-metadata", computer_name="PC-META")
    db_session.add_all([printer, agent])
    db_session.flush()
    alias = PrinterAlias(
        organization_id=1,
        printer=printer,
        agent=agent,
        queue_name="KONICA Financeiro",
        normalized_queue_name="konica financeiro",
        computer_name="PC-META",
        driver_name="KONICA Driver",
        port_name="IP_192.168.1.125",
        connection_type="network",
        ip_address="192.168.1.125",
        serial_number="SN-META-001",
        device_id="MFG:KONICA;MDL:C368;",
        fingerprint="serial:sn-meta-001",
    )
    db_session.add(alias)
    db_session.commit()

    register_print_job(
        db_session,
        PrintJobCreate(
            username="diego",
            printer_name="KONICA Financeiro",
            queue_name="  konica   financeiro ",
            pages=1,
            is_color=False,
            agent_uid=agent.agent_uid,
            external_job_id="eventlog:metadata-preserved",
        ),
    )

    db_session.refresh(alias)
    assert alias.printer_id == printer.id
    assert alias.computer_name == "PC-META"
    assert alias.driver_name == "KONICA Driver"
    assert alias.port_name == "IP_192.168.1.125"
    assert alias.connection_type == "network"
    assert alias.ip_address == "192.168.1.125"
    assert alias.serial_number == "SN-META-001"
    assert alias.device_id == "MFG:KONICA;MDL:C368;"
    assert alias.fingerprint == "serial:sn-meta-001"


def test_usb_job_uses_known_alias_device_id_instead_of_creating_queue_printer(db_session: Session):
    printer = Printer(organization_id=1, name="BROTHER DCP-T420W", is_color=False)
    alias = PrinterAlias(
        organization_id=1,
        printer=printer,
        queue_name="Brother USB",
        connection_type="usb",
        device_id="USBPRINT\\BROTHERDCP-T420W\\7&ABC",
    )
    db_session.add_all([printer, alias])
    db_session.commit()

    register_print_job(
        db_session,
        PrintJobCreate(
            username="diego",
            printer_name="USER",
            pages=2,
            is_color=False,
            external_job_id="eventlog:usb-device-id",
            agent_uid="agent-usb-job",
            computer_name="PC-USB",
            queue_name="USER",
            printer_port_name="USB001",
            printer_driver_name="Brother Driver",
            printer_connection_type="usb",
            printer_device_id="USBPRINT\\BROTHERDCP-T420W\\7&ABC",
            printer_fingerprint="usb:pc-usb|usbprint\\brotherdcp-t420w\\7&abc",
        ),
    )

    assert db_session.query(Printer).count() == 1
    job = db_session.query(PrintJob).one()
    assert job.printer_id == printer.id
    assert job.printer_alias_id != alias.id
    assert job.queue_name == "USER"
    user_alias = db_session.query(PrinterAlias).filter(PrinterAlias.queue_name == "USER").one()
    assert user_alias.printer_id == printer.id


def test_usb_job_matches_known_alias_device_id_case_insensitive(db_session: Session):
    printer = Printer(organization_id=1, name="BROTHER USB CASE", is_color=False)
    alias = PrinterAlias(
        organization_id=1,
        printer=printer,
        queue_name="Brother USB Original",
        connection_type="usb",
        device_id="USBPRINT\\BROTHERDCP-T420W\\7&ABC",
    )
    db_session.add_all([printer, alias])
    db_session.commit()

    register_print_job(
        db_session,
        PrintJobCreate(
            username="diego",
            printer_name="USER",
            pages=1,
            is_color=False,
            external_job_id="eventlog:usb-device-id-case",
            agent_uid="agent-usb-case",
            computer_name="PC-USB-CASE",
            queue_name="USER",
            printer_port_name="USB001",
            printer_driver_name="Brother Driver",
            printer_connection_type="usb",
            printer_device_id="usbprint\\brotherdcp-t420w\\7&abc",
            printer_fingerprint="usb:pc-usb-case|usbprint\\brotherdcp-t420w\\7&abc",
        ),
    )

    job = db_session.query(PrintJob).one()
    assert db_session.query(Printer).count() == 1
    assert job.printer_id == printer.id
    assert db_session.query(PrinterAlias).filter(PrinterAlias.queue_name == "USER").one().printer_id == printer.id


def test_network_job_uses_known_fingerprint_instead_of_creating_duplicate_printer(db_session: Session):
    printer = Printer(organization_id=1, name="KONICA FISICA WSD", is_color=True)
    known_alias = PrinterAlias(
        organization_id=1,
        printer=printer,
        queue_name="KONICA Financeiro",
        connection_type="network",
        port_name="WSD-12345",
        driver_name="KONICA Driver",
        fingerprint="network:wsd-12345|konica driver",
    )
    db_session.add_all([printer, known_alias])
    db_session.commit()

    register_print_job(
        db_session,
        PrintJobCreate(
            username="diego",
            printer_name="Impressora Sala",
            pages=3,
            is_color=True,
            external_job_id="eventlog:network-fingerprint",
            agent_uid="agent-network-fingerprint",
            computer_name="PC-RH",
            queue_name="Impressora Sala",
            printer_port_name="WSD-12345",
            printer_driver_name="KONICA Driver",
            printer_connection_type="network",
            printer_fingerprint="network:wsd-12345|konica driver",
        ),
    )

    job = db_session.query(PrintJob).one()
    local_alias = db_session.query(PrinterAlias).filter(PrinterAlias.queue_name == "Impressora Sala").one()
    assert db_session.query(Printer).count() == 1
    assert job.printer_id == printer.id
    assert local_alias.printer_id == printer.id
    assert local_alias.fingerprint == "network:wsd-12345|konica driver"


def test_generic_printer_name_uses_single_bound_agent_printer(db_session: Session):
    printer = Printer(organization_id=1, name="KONICA MINOLTA C368SeriesPS", is_color=True)
    agent = PrintAgent(organization_id=1, agent_uid="agent-single-printer", computer_name="PC-FIN")
    db_session.add_all([printer, agent])
    db_session.flush()
    db_session.add(
        PrinterAlias(
            organization_id=1,
            printer=printer,
            agent=agent,
            queue_name="KONICA Financeiro",
            normalized_queue_name="konica financeiro",
            driver_name="KONICA Driver",
        )
    )
    db_session.commit()

    register_print_job(
        db_session,
        PrintJobCreate(
            username="diego",
            printer_name="USER",
            queue_name="USER",
            pages=1,
            is_color=False,
            external_job_id="eventlog:generic-single-printer",
            agent_uid=agent.agent_uid,
            computer_name=agent.computer_name,
        ),
    )

    job = db_session.query(PrintJob).one()
    assert db_session.query(Printer).count() == 1
    assert job.printer_id == printer.id
    assert db_session.query(PrinterAlias).filter(PrinterAlias.queue_name == "USER").one().printer_id == printer.id


def test_generic_printer_name_does_not_guess_when_agent_has_multiple_bound_printers(db_session: Session):
    konica = Printer(organization_id=1, name="KONICA FINANCEIRO", is_color=True)
    brother = Printer(organization_id=1, name="BROTHER RH", is_color=False)
    agent = PrintAgent(organization_id=1, agent_uid="agent-multiple-printers", computer_name="PC-MULTI")
    db_session.add_all([konica, brother, agent])
    db_session.flush()
    db_session.add_all(
        [
            PrinterAlias(
                organization_id=1,
                printer=konica,
                agent=agent,
                queue_name="KONICA Financeiro",
                normalized_queue_name="konica financeiro",
            ),
            PrinterAlias(
                organization_id=1,
                printer=brother,
                agent=agent,
                queue_name="Brother RH",
                normalized_queue_name="brother rh",
            ),
        ]
    )
    db_session.commit()

    register_print_job(
        db_session,
        PrintJobCreate(
            username="diego",
            printer_name="USER",
            queue_name="USER",
            pages=1,
            is_color=False,
            external_job_id="eventlog:generic-multiple-printers",
            agent_uid=agent.agent_uid,
            computer_name=agent.computer_name,
        ),
    )

    job = db_session.query(PrintJob).filter(PrintJob.external_job_id == "eventlog:generic-multiple-printers").one()
    user_printer = db_session.query(Printer).filter(Printer.name == "USER").one()
    assert db_session.query(Printer).count() == 3
    assert job.printer_id == user_printer.id


def test_merge_printer_moves_jobs_and_aliases(db_session: Session):
    actor = User(username="admin", full_name="Admin", role=UserRole.admin, is_active=True)
    user = User(username="diego", full_name="Diego", role=UserRole.user, is_active=True)
    target = Printer(name="KONICA MINOLTA C368SeriesPS", ip_address="192.168.1.125", is_color=True)
    source = Printer(name="USER", is_color=False)
    agent = PrintAgent(agent_uid="agent-pc-a", computer_name="PC-A")
    db_session.add_all([actor, user, target, source, agent])
    db_session.flush()

    alias = PrinterAlias(
        printer_id=source.id,
        agent_id=agent.id,
        queue_name="USER",
        normalized_queue_name="user",
        computer_name="PC-A",
        fingerprint="queue:pc-a|user|usb001|driver",
    )
    db_session.add(alias)
    db_session.flush()
    job = PrintJob(
        user_id=user.id,
        printer_id=source.id,
        printer_alias_id=alias.id,
        agent_id=agent.id,
        document_name="Teste",
        computer_name="PC-A",
        queue_name="USER",
        pages=1,
        is_color=False,
        cost=0.05,
        status=JobStatus.authorized,
        submitted_at=datetime.now(timezone.utc),
    )
    db_session.add(job)
    db_session.commit()

    merge_printer_endpoint(source.id, target.id, db_session, actor)

    moved_job = db_session.query(PrintJob).filter(PrintJob.id == job.id).one()
    moved_alias = db_session.query(PrinterAlias).filter(PrinterAlias.id == alias.id).one()
    assert db_session.query(Printer).filter(Printer.id == source.id).first() is None
    assert moved_job.printer_id == target.id
    assert moved_job.printer_alias_id == alias.id
    assert moved_alias.printer_id == target.id


def test_merge_printer_collapses_duplicate_aliases_by_normalized_queue_name(db_session: Session):
    actor = User(username="admin-merge-normalized", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    user = User(username="diego-merge-normalized", full_name="Diego", role=UserRole.user, is_active=True, organization_id=1)
    target = Printer(organization_id=1, name="KONICA FINANCEIRO", ip_address="192.168.1.125", is_color=True)
    source = Printer(organization_id=1, name="KONICA DUPLICADA", is_color=True)
    agent = PrintAgent(organization_id=1, agent_uid="agent-merge-normalized", computer_name="PC-FIN")
    db_session.add_all([actor, user, target, source, agent])
    db_session.flush()

    target_alias = PrinterAlias(
        organization_id=1,
        printer_id=target.id,
        agent_id=agent.id,
        queue_name="KONICA Financeiro",
        normalized_queue_name="konica financeiro",
        computer_name="PC-FIN",
    )
    source_alias = PrinterAlias(
        organization_id=1,
        printer_id=source.id,
        agent_id=agent.id,
        queue_name="  konica   financeiro ",
        normalized_queue_name="konica financeiro",
        computer_name="PC-FIN",
    )
    db_session.add_all([target_alias, source_alias])
    db_session.flush()
    job = PrintJob(
        organization_id=1,
        user_id=user.id,
        printer_id=source.id,
        printer_alias_id=source_alias.id,
        agent_id=agent.id,
        document_name="Contrato.pdf",
        computer_name="PC-FIN",
        queue_name=source_alias.queue_name,
        pages=3,
        is_color=True,
        cost=0.75,
        status=JobStatus.authorized,
        submitted_at=datetime.now(timezone.utc),
    )
    db_session.add(job)
    db_session.commit()

    merge_printer_endpoint(source.id, target.id, db_session, actor)

    moved_job = db_session.query(PrintJob).filter(PrintJob.id == job.id).one()
    aliases = db_session.query(PrinterAlias).filter(PrinterAlias.printer_id == target.id).all()
    audit = db_session.query(AuditLog).filter(AuditLog.action == "printer_merged").one()
    assert db_session.query(Printer).filter(Printer.id == source.id).first() is None
    assert moved_job.printer_id == target.id
    assert moved_job.printer_alias_id == target_alias.id
    assert [alias.id for alias in aliases] == [target_alias.id]
    assert audit.log_metadata["merged_aliases"] == 1


def test_merge_printer_preserves_printer_and_alias_policies(db_session: Session):
    actor = User(username="admin-merge-policies", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    target = Printer(organization_id=1, name="KONICA POLITICAS", ip_address="192.168.1.125", is_color=True)
    source = Printer(organization_id=1, name="KONICA POLITICAS DUP", is_color=True)
    agent = PrintAgent(organization_id=1, agent_uid="agent-merge-policies", computer_name="PC-POL")
    db_session.add_all([actor, target, source, agent])
    db_session.flush()

    target_alias = PrinterAlias(
        organization_id=1,
        printer_id=target.id,
        agent_id=agent.id,
        queue_name="KONICA Financeiro",
        normalized_queue_name="konica financeiro",
        computer_name="PC-POL",
    )
    source_alias = PrinterAlias(
        organization_id=1,
        printer_id=source.id,
        agent_id=agent.id,
        queue_name="  konica   financeiro ",
        normalized_queue_name="konica financeiro",
        computer_name="PC-POL",
    )
    db_session.add_all([target_alias, source_alias])
    db_session.flush()

    printer_policy = PrintPolicy(
        organization_id=1,
        name="Bloquear colorido na duplicada",
        priority=10,
        rule_type=PolicyRuleType.color,
        action=PolicyAction.block,
        printer_id=source.id,
    )
    alias_policy = PrintPolicy(
        organization_id=1,
        name="Liberar fila duplicada acima de 20",
        priority=20,
        rule_type=PolicyRuleType.max_pages,
        action=PolicyAction.require_release,
        max_pages=20,
        printer_alias_id=source_alias.id,
    )
    db_session.add_all([printer_policy, alias_policy])
    db_session.commit()

    merge_printer_endpoint(source.id, target.id, db_session, actor)

    db_session.refresh(printer_policy)
    db_session.refresh(alias_policy)
    audit = db_session.query(AuditLog).filter(AuditLog.action == "printer_merged").one()
    assert printer_policy.printer_id == target.id
    assert alias_policy.printer_alias_id == target_alias.id
    assert db_session.query(PrinterAlias).filter(PrinterAlias.id == source_alias.id).first() is None
    assert audit.log_metadata["moved_policies"] == 1
    assert audit.log_metadata["moved_alias_policies"] == 1


def test_merge_printer_moves_queue_actions_to_target_printer(db_session: Session):
    actor = User(username="merge-queue-action-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    source = Printer(organization_id=1, name="KONICA DUPLICADA", is_color=True)
    target = Printer(organization_id=1, name="KONICA FINAL", is_color=True)
    agent = PrintAgent(organization_id=1, agent_uid="agent-merge-queue-action", computer_name="PC-MERGE-ACTION")
    db_session.add_all([actor, source, target, agent])
    db_session.flush()
    action = AgentQueueAction(
        organization_id=1,
        agent_id=agent.id,
        printer_id=source.id,
        requested_by_user_id=actor.id,
        action_type=AgentQueueActionType.restore_queue,
        queue_name="KONICA Financeiro",
        driver_name="KONICA Driver",
        ip_address="192.168.1.125",
        status=AgentQueueActionStatus.pending,
    )
    db_session.add(action)
    db_session.commit()

    merge_printer_endpoint(source.id, target.id, db_session, actor)
    db_session.refresh(action)
    audit = db_session.query(AuditLog).filter(AuditLog.action == "printer_merged", AuditLog.entity_id == target.id).one()

    assert action.printer_id == target.id
    assert db_session.get(Printer, source.id) is None
    assert audit.log_metadata["moved_queue_actions"] == 1


def test_printer_with_linked_policies_cannot_be_deleted(db_session: Session):
    actor = User(username="admin-delete-policy-printer", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    printer = Printer(organization_id=1, name="KONICA COM POLITICA", is_color=True)
    agent = PrintAgent(organization_id=1, agent_uid="agent-delete-policy-printer", computer_name="PC-POL")
    db_session.add_all([actor, printer, agent])
    db_session.flush()

    alias = PrinterAlias(
        organization_id=1,
        printer_id=printer.id,
        agent_id=agent.id,
        queue_name="KONICA Politica",
        normalized_queue_name="konica politica",
    )
    db_session.add(alias)
    db_session.flush()
    db_session.add_all(
        [
            PrintPolicy(
                organization_id=1,
                name="Bloquear colorido impressora",
                priority=10,
                rule_type=PolicyRuleType.color,
                action=PolicyAction.block,
                printer_id=printer.id,
            ),
            PrintPolicy(
                organization_id=1,
                name="Liberar alias impressora",
                priority=20,
                rule_type=PolicyRuleType.max_pages,
                action=PolicyAction.require_release,
                max_pages=20,
                printer_alias_id=alias.id,
            ),
        ]
    )
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        delete_printer_endpoint(printer.id, db=db_session, actor=actor)

    assert exc.value.status_code == 409
    assert "politicas vinculadas" in exc.value.detail
    assert db_session.get(Printer, printer.id) is not None
    assert db_session.get(PrinterAlias, alias.id) is not None
    assert db_session.query(PrintPolicy).filter(PrintPolicy.organization_id == 1).count() == 2


def test_printer_with_pending_queue_action_cannot_be_deleted(db_session: Session):
    actor = User(username="admin-delete-queue-printer", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    printer = Printer(organization_id=1, name="KONICA COM ACAO", is_color=True)
    agent = PrintAgent(organization_id=1, agent_uid="agent-delete-queue-printer", computer_name="PC-QUEUE")
    db_session.add_all([actor, printer, agent])
    db_session.flush()
    action = AgentQueueAction(
        organization_id=1,
        agent_id=agent.id,
        printer_id=printer.id,
        requested_by_user_id=actor.id,
        action_type=AgentQueueActionType.create_queue,
        queue_name="KONICA Financeiro",
        driver_name="KONICA Driver",
        ip_address="192.168.1.125",
        status=AgentQueueActionStatus.pending,
    )
    db_session.add(action)
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        delete_printer_endpoint(printer.id, db=db_session, actor=actor)

    assert exc.value.status_code == 409
    assert "acoes remotas pendentes" in exc.value.detail
    assert db_session.get(Printer, printer.id) is not None
    assert db_session.get(AgentQueueAction, action.id).printer_id == printer.id


def test_delete_printer_detaches_completed_queue_actions(db_session: Session):
    actor = User(username="admin-delete-done-queue-printer", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    printer = Printer(organization_id=1, name="KONICA ACAO FINALIZADA", is_color=True)
    agent = PrintAgent(organization_id=1, agent_uid="agent-delete-done-queue-printer", computer_name="PC-QUEUE-DONE")
    db_session.add_all([actor, printer, agent])
    db_session.flush()
    action = AgentQueueAction(
        organization_id=1,
        agent_id=agent.id,
        printer_id=printer.id,
        requested_by_user_id=actor.id,
        action_type=AgentQueueActionType.restore_queue,
        queue_name="KONICA Financeiro",
        driver_name="KONICA Driver",
        ip_address="192.168.1.125",
        status=AgentQueueActionStatus.succeeded,
        result_message="Fila restaurada",
    )
    db_session.add(action)
    db_session.commit()

    result = delete_printer_endpoint(printer.id, db=db_session, actor=actor)
    db_session.refresh(action)
    audit = db_session.query(AuditLog).filter(AuditLog.action == "printer_deleted", AuditLog.entity_id == printer.id).one()

    assert result == {"status": "deleted", "deleted_jobs": 0, "detached_queue_actions": 1}
    assert action.printer_id is None
    assert db_session.get(Printer, printer.id) is None
    assert audit.log_metadata["detached_queue_actions"] == 1


def test_binding_alias_moves_historical_jobs_to_physical_printer(db_session: Session):
    actor = User(username="admin-bind", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    user = User(username="diego-bind", full_name="Diego", role=UserRole.user, is_active=True, organization_id=1)
    detected = Printer(organization_id=1, name="USER", is_color=False)
    physical = Printer(organization_id=1, name="KONICA MINOLTA C368SeriesPS", ip_address="192.168.1.125", is_color=True)
    agent = PrintAgent(organization_id=1, agent_uid="agent-bind", computer_name="PC-A")
    db_session.add_all([actor, user, detected, physical, agent])
    db_session.flush()

    alias = PrinterAlias(
        organization_id=1,
        printer_id=detected.id,
        agent_id=agent.id,
        queue_name="USER",
        normalized_queue_name="user",
        computer_name="PC-A",
        fingerprint="queue:pc-a|user|usb001|driver",
    )
    db_session.add(alias)
    db_session.flush()
    job = PrintJob(
        organization_id=1,
        user_id=user.id,
        printer_id=detected.id,
        printer_alias_id=alias.id,
        agent_id=agent.id,
        document_name="Documento",
        computer_name="PC-A",
        queue_name="USER",
        pages=2,
        is_color=False,
        cost=0.10,
        status=JobStatus.authorized,
        submitted_at=datetime.now(timezone.utc),
    )
    db_session.add(job)
    db_session.commit()

    updated_alias = bind_printer_alias_endpoint(alias.id, PrinterAliasBind(printer_id=physical.id), db_session, actor)

    moved_job = db_session.query(PrintJob).filter(PrintJob.id == job.id).one()
    assert updated_alias.printer_id == physical.id
    assert moved_job.printer_id == physical.id


def test_binding_alias_is_scoped_by_organization(db_session: Session):
    other_org = Organization(name="Cliente Alias B", slug="cliente-alias-b", is_active=True)
    db_session.add(other_org)
    db_session.flush()
    actor = User(username="admin-alias-a", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    alias = PrinterAlias(organization_id=1, queue_name="KONICA LOCAL")
    other_printer = Printer(organization_id=other_org.id, name="KONICA OUTRA EMPRESA", is_color=True)
    db_session.add_all([actor, alias, other_printer])
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        bind_printer_alias_endpoint(alias.id, PrinterAliasBind(printer_id=other_printer.id), db_session, actor)
    assert exc.value.status_code == 404

    db_session.rollback()
    unchanged_alias = db_session.query(PrinterAlias).filter(PrinterAlias.id == alias.id).one()
    assert unchanged_alias.printer_id is None


def test_agent_printer_status_update_requires_bound_alias(db_session: Session):
    admin = User(username="admin-status", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    agent_user = User(username="agent-status", full_name="Agent", role=UserRole.agent, is_active=True, organization_id=1)
    agent = PrintAgent(organization_id=1, agent_uid="agent-status-pc", computer_name="PC-STATUS")
    bound_printer = Printer(organization_id=1, name="KONICA STATUS", ip_address="192.168.1.125", is_color=True)
    other_printer = Printer(organization_id=1, name="BROTHER STATUS", ip_address="192.168.1.126", is_color=False)
    db_session.add_all([admin, agent_user, agent, bound_printer, other_printer])
    db_session.flush()
    db_session.add(
        PrinterAlias(
            organization_id=1,
            printer_id=bound_printer.id,
            agent_id=agent.id,
            queue_name="KONICA LOCAL",
            normalized_queue_name="konica local",
        )
    )
    db_session.commit()

    updated = update_printer_status_endpoint(
        bound_printer.id,
        PrinterStatusUpdate(agent_uid=agent.agent_uid, page_counter=1234, toner_level=88),
        db_session,
        agent_user,
    )

    assert updated.page_counter == 1234
    assert updated.toner_level == 88

    with pytest.raises(HTTPException) as missing_uid:
        update_printer_status_endpoint(
            bound_printer.id,
            PrinterStatusUpdate(page_counter=2222),
            db_session,
            agent_user,
        )
    assert missing_uid.value.status_code == 403

    with pytest.raises(HTTPException) as unbound_printer:
        update_printer_status_endpoint(
            other_printer.id,
            PrinterStatusUpdate(agent_uid=agent.agent_uid, page_counter=3333),
            db_session,
            agent_user,
        )
    assert unbound_printer.value.status_code == 403

    admin_updated = update_printer_status_endpoint(
        other_printer.id,
        PrinterStatusUpdate(page_counter=4444),
        db_session,
        admin,
    )
    assert admin_updated.page_counter == 4444


def test_organization_scope_isolates_core_views(db_session: Session, monkeypatch, tmp_path: Path):
    monkeypatch.setattr(settings, "agent_download_dir", str(tmp_path))
    monkeypatch.setattr(settings, "agent_latest_version", "0.9.0")
    monkeypatch.setattr(settings, "agent_download_filename", "PrintBillingAgent.exe")
    (tmp_path / "PrintBillingAgent.exe").write_bytes(b"agent-v9")

    default_org = db_session.query(Organization).filter(Organization.id == 1).one()
    default_org.billing_plan = "professional"
    default_org.billing_status = "active"
    default_org.contracted_printer_limit = 3
    other_org = Organization(name="Cliente B", slug="cliente-b", is_active=True)
    db_session.add(other_org)
    db_session.flush()

    org_one_department = Department(organization_id=1, name="Financeiro", cost_center="CC-FIN")
    org_one_admin = User(username="org1-admin", full_name="Org 1 Admin", role=UserRole.admin, is_active=True, organization_id=1)
    org_one_user = User(username="org1-user", full_name="Org 1 User", role=UserRole.user, is_active=True, organization_id=1, department=org_one_department)
    org_two_user = User(username="org2-user", full_name="Org 2 User", role=UserRole.user, is_active=True, organization_id=other_org.id)
    org_one_printer = Printer(
        name="Org 1 Printer",
        is_color=False,
        organization_id=1,
        ip_address="192.168.10.20",
        toner_levels={"black": 8},
    )
    org_one_duplicate_printer = Printer(
        name="Org 1 Duplicate Printer",
        is_color=False,
        organization_id=1,
        ip_address="192.168.10.20",
    )
    org_two_printer = Printer(name="Org 2 Printer", is_color=False, organization_id=other_org.id, ip_address="192.168.20.20")
    db_session.add_all([org_one_department, org_one_admin, org_one_user, org_two_user, org_one_printer, org_one_duplicate_printer, org_two_printer])
    db_session.flush()
    now = datetime.now(timezone.utc)
    org_one_online_agent = PrintAgent(
        organization_id=1,
        agent_uid="org1-online-agent",
        computer_name="PC-ONLINE",
        event_log_enabled=True,
        last_seen_at=now,
    )
    org_one_alert_agent = PrintAgent(
        organization_id=1,
        agent_uid="org1-alert-agent",
        computer_name="PC-ALERT",
        event_log_enabled=False,
        local_admin=False,
        last_error="Event Log indisponivel",
        last_seen_at=now - timedelta(minutes=10),
    )
    org_one_stale_action_agent = PrintAgent(
        organization_id=1,
        agent_uid="org1-stale-action-agent",
        computer_name="PC-STALE-ACTION",
        event_log_enabled=True,
        last_seen_at=now,
    )
    org_one_recent_log_agent = PrintAgent(
        organization_id=1,
        agent_uid="org1-recent-log-agent",
        computer_name="PC-RECENT-LOG",
        event_log_enabled=True,
        last_seen_at=now,
    )
    org_one_outdated_agent = PrintAgent(
        organization_id=1,
        agent_uid="org1-outdated-agent",
        computer_name="PC-OUTDATED",
        event_log_enabled=True,
        version="0.1.0",
        last_seen_at=now - timedelta(seconds=130),
    )
    org_two_agent = PrintAgent(
        organization_id=other_org.id,
        agent_uid="org2-online-agent",
        computer_name="PC-ORG2",
        event_log_enabled=True,
        last_seen_at=now,
    )
    db_session.add_all([org_one_online_agent, org_one_alert_agent, org_one_stale_action_agent, org_one_recent_log_agent, org_one_outdated_agent, org_two_agent])
    db_session.flush()

    db_session.add_all(
        [
            PrinterAlias(
                organization_id=1,
                agent_id=org_one_online_agent.id,
                queue_name="Org 1 USB sem vinculo",
                connection_type="usb",
                last_seen_at=now,
            ),
            PrinterAlias(
                organization_id=1,
                agent_id=org_one_online_agent.id,
                queue_name="Org 1 fila antiga sem vinculo",
                connection_type="network",
                last_seen_at=now - timedelta(minutes=30),
            ),
            PrinterAlias(
                organization_id=1,
                printer_id=org_one_printer.id,
                agent_id=org_one_online_agent.id,
                queue_name="Org 1 Printer",
                normalized_queue_name="org 1 printer",
                connection_type="network",
                ip_address="192.168.10.20",
                last_seen_at=now,
            ),
            PrinterAlias(
                organization_id=1,
                printer_id=org_one_printer.id,
                agent_id=org_one_online_agent.id,
                queue_name="  org   1   printer ",
                connection_type="network",
                last_seen_at=now,
            ),
            PrinterAlias(
                organization_id=1,
                printer_id=org_one_printer.id,
                agent_id=org_one_stale_action_agent.id,
                queue_name="Org 1 Stale Action Printer",
                normalized_queue_name="org 1 stale action printer",
                connection_type="network",
                last_seen_at=now,
            ),
            AgentQueueAction(
                organization_id=1,
                agent_id=org_one_stale_action_agent.id,
                action_type=AgentQueueActionType.create_queue,
                queue_name="Org 1 Stale Action Printer",
                status=AgentQueueActionStatus.pending,
                requested_at=now - timedelta(minutes=20),
            ),
            PrinterAlias(
                organization_id=1,
                printer_id=org_one_printer.id,
                agent_id=org_one_recent_log_agent.id,
                queue_name="Org 1 Recent Log Printer",
                normalized_queue_name="org 1 recent log printer",
                connection_type="network",
                last_seen_at=now,
            ),
            PrinterAlias(
                organization_id=1,
                printer_id=org_one_printer.id,
                agent_id=org_one_recent_log_agent.id,
                queue_name="USER",
                normalized_queue_name="user",
                connection_type="network",
                last_seen_at=now,
            ),
            PrinterAlias(
                organization_id=1,
                printer_id=org_one_duplicate_printer.id,
                agent_id=org_one_recent_log_agent.id,
                queue_name="Org 1 Duplicate Printer",
                normalized_queue_name="org 1 duplicate printer",
                connection_type="network",
                ip_address="192.168.10.20",
                last_seen_at=now,
            ),
            AgentLog(
                organization_id=1,
                agent_id=org_one_recent_log_agent.id,
                level="error",
                message="Falha recente ao consultar fila local",
                source="spool",
                occurred_at=now - timedelta(minutes=1),
                received_at=now - timedelta(minutes=1),
            ),
            PrinterAlias(
                organization_id=1,
                printer_id=org_one_printer.id,
                agent_id=org_one_outdated_agent.id,
                queue_name="Org 1 Outdated Printer",
                normalized_queue_name="org 1 outdated printer",
                connection_type="network",
                last_seen_at=now,
            ),
            PrinterAlias(
                organization_id=other_org.id,
                agent_id=org_two_agent.id,
                queue_name="Org 2 sem vinculo",
                connection_type="usb",
                last_seen_at=now,
            ),
            PrintJob(
                organization_id=1,
                user_id=org_one_user.id,
                printer_id=org_one_printer.id,
                pages=3,
                is_color=False,
                cost=0.15,
                status=JobStatus.authorized,
                submitted_at=datetime.now(timezone.utc),
            ),
            PrintJob(
                organization_id=1,
                user_id=org_two_user.id,
                printer_id=org_two_printer.id,
                pages=99,
                is_color=False,
                cost=9.90,
                status=JobStatus.authorized,
                submitted_at=datetime.now(timezone.utc),
            ),
            PrintJob(
                organization_id=other_org.id,
                user_id=org_two_user.id,
                printer_id=org_two_printer.id,
                pages=7,
                is_color=False,
                cost=0.35,
                status=JobStatus.authorized,
                submitted_at=datetime.now(timezone.utc),
            ),
        ]
    )
    db_session.commit()

    users = list_users(db=db_session, _=org_one_admin)
    departments = list_departments(db=db_session, actor=org_one_admin)
    printers = list_printers(db=db_session, actor=org_one_admin)
    jobs = list_jobs(
        user_id=None,
        department_id=None,
        printer_id=None,
        date_from=None,
        date_to=None,
        db=db_session,
        actor=org_one_admin,
    )
    cost_center_jobs = list_jobs(
        user_id=None,
        department_id=None,
        cost_center="CC-FIN",
        printer_id=None,
        date_from=None,
        date_to=None,
        db=db_session,
        actor=org_one_admin,
    )
    metrics = dashboard_metrics(db_session, organization_id=1)
    agents = list_agents(db=db_session, actor=org_one_admin)
    export_response = export_report(format="xlsx", db=db_session, actor=org_one_admin)
    workbook = load_workbook(BytesIO(export_response.body), data_only=True)
    exported_users = [row[1].value for row in workbook["Impressoes"].iter_rows(min_row=2)]
    snapshot = build_monthly_snapshot(db_session, organization_id=1, year=datetime.now(timezone.utc).year, month=datetime.now(timezone.utc).month)

    assert {user.username for user in users} == {"org1-admin", "org1-user"}
    assert [department.name for department in departments] == ["Financeiro"]
    org_user_read = next(user for user in users if user.username == "org1-user")
    assert org_user_read.department_id == org_one_department.id
    assert org_user_read.department_name == "Financeiro"
    assert [printer.name for printer in printers] == ["Org 1 Duplicate Printer", "Org 1 Printer"]
    conflict_by_name = {printer.name: printer for printer in printers}
    assert conflict_by_name["Org 1 Duplicate Printer"].identity_conflict_count == 1
    assert conflict_by_name["Org 1 Duplicate Printer"].identity_conflict_types == ["ip"]
    assert conflict_by_name["Org 1 Duplicate Printer"].identity_conflict_printer_ids == [org_one_printer.id]
    assert conflict_by_name["Org 1 Printer"].identity_conflict_count == 1
    assert conflict_by_name["Org 1 Printer"].identity_conflict_printer_ids == [org_one_duplicate_printer.id]
    assert [job.username for job in jobs] == ["org1-user"]
    assert [job.username for job in cost_center_jobs] == ["org1-user"]
    assert jobs[0].department_id == org_one_department.id
    assert jobs[0].department_name == "Financeiro"
    assert jobs[0].department_cost_center == "CC-FIN"
    assert jobs[0].cost == 0.15
    assert metrics["pages_month"] == 3
    assert metrics["contract_overview"] == {
        "billing_plan": "professional",
        "billing_status": "active",
        "contracted_printer_limit": 3,
        "active_printers_count": 2,
        "printer_usage_percent": 66.7,
        "printer_limit_status": "ok",
    }
    assert metrics["top_users"] == [
        {"username": "Org 1 User", "pages": 3, "cost": 0.15, "cost_per_page": 0.05},
    ]
    assert metrics["top_printers"] == [
        {"printer": "Org 1 Printer", "pages": 3, "cost": 0.15, "cost_per_page": 0.05},
    ]
    assert metrics["operational_health"] == {
        "agents_total": 5,
        "agents_online": 4,
        "agents_offline": 1,
        "agents_with_alerts": 5,
        "agents_with_delayed_heartbeat": 1,
        "agents_without_local_admin": 1,
        "agents_without_event_log": 1,
        "outdated_agents": 5,
        "agents_with_recent_errors": 1,
        "printers_total": 2,
        "printers_monitored": 2,
        "printers_unmonitored": 0,
        "low_toner_printers": 1,
        "unbound_queues": 1,
        "usb_queues": 1,
        "duplicate_queue_aliases": 1,
        "generic_queue_aliases": 1,
        "hardware_identity_conflicts": 1,
        "pending_queue_actions": 1,
        "stale_queue_actions": 1,
    }
    validated_metrics = DashboardMetrics.model_validate(metrics)
    assert validated_metrics.contract_overview is not None
    assert validated_metrics.contract_overview.printer_usage_percent == 66.7
    assert validated_metrics.operational_health is not None
    assert validated_metrics.operational_health.agents_without_local_admin == 1
    assert validated_metrics.operational_health.agents_with_delayed_heartbeat == 1
    assert validated_metrics.operational_health.agents_without_event_log == 1
    assert validated_metrics.operational_health.outdated_agents == 5
    assert validated_metrics.operational_health.agents_with_recent_errors == 1
    assert validated_metrics.operational_health.duplicate_queue_aliases == 1
    assert validated_metrics.operational_health.generic_queue_aliases == 1
    assert validated_metrics.operational_health.hardware_identity_conflicts == 1
    assert validated_metrics.operational_health.pending_queue_actions == 1
    assert validated_metrics.operational_health.stale_queue_actions == 1
    recent_log_agent = next(agent for agent in agents if agent.agent_uid == "org1-recent-log-agent")
    assert any(alert.code == "hardware_identity_conflict" for alert in recent_log_agent.health_alerts)
    assert validated_metrics.top_users[0].username == "Org 1 User"
    assert validated_metrics.top_printers[0].printer == "Org 1 Printer"
    assert validated_metrics.department_usage[0].department == "Financeiro"
    assert validated_metrics.cost_center_usage[0].cost_center == "CC-FIN"
    assert validated_metrics.eco_metrics is not None
    assert validated_metrics.eco_metrics.pages_saved == 0
    assert metrics["department_usage"] == [
        {"department": "Financeiro", "pages": 3, "cost": 0.15, "cost_per_page": 0.05},
    ]
    assert metrics["cost_center_usage"] == [
        {"cost_center": "CC-FIN", "pages": 3, "cost": 0.15, "cost_per_page": 0.05},
    ]
    assert exported_users == ["Org 1 User"]
    assert snapshot["totals"]["total_jobs"] == 1
    assert snapshot["totals"]["total_pages"] == 3
    assert snapshot["by_user"] == [
        {
            "name": "Org 1 User",
            "jobs": 1,
            "pages": 3,
            "mono_pages": 3,
            "color_pages": 0,
            "cost": 0.15,
            "cost_per_page": 0.05,
            "page_share_percent": 100.0,
            "cost_share_percent": 100.0,
        },
    ]
    assert snapshot["by_cost_center"] == [
        {
            "name": "CC-FIN",
            "jobs": 1,
            "pages": 3,
            "mono_pages": 3,
            "color_pages": 0,
            "cost": 0.15,
            "cost_per_page": 0.05,
            "page_share_percent": 100.0,
            "cost_share_percent": 100.0,
        },
    ]


def test_job_listing_rejects_filters_from_other_organization(db_session: Session):
    other_org = Organization(name="Cliente Jobs B", slug="cliente-jobs-b", is_active=True)
    actor = User(username="jobs-scope-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    other_department = Department(organization=other_org, name="Financeiro", cost_center="CC-OUTRO")
    other_user = User(username="jobs-scope-user", full_name="Outro Usuario", role=UserRole.user, is_active=True, organization=other_org, department=other_department)
    other_printer = Printer(organization=other_org, name="KONICA OUTRA JOBS", is_color=True)
    db_session.add_all([other_org, actor, other_department, other_user, other_printer])
    db_session.commit()

    with pytest.raises(HTTPException) as user_exc:
        list_jobs(user_id=other_user.id, department_id=None, printer_id=None, date_from=None, date_to=None, db=db_session, actor=actor)
    assert user_exc.value.status_code == 404

    with pytest.raises(HTTPException) as department_exc:
        list_jobs(user_id=None, department_id=other_department.id, printer_id=None, date_from=None, date_to=None, db=db_session, actor=actor)
    assert department_exc.value.status_code == 404

    with pytest.raises(HTTPException) as cost_center_exc:
        list_jobs(user_id=None, department_id=None, cost_center="CC-OUTRO", printer_id=None, date_from=None, date_to=None, db=db_session, actor=actor)
    assert cost_center_exc.value.status_code == 404

    with pytest.raises(HTTPException) as printer_exc:
        list_jobs(user_id=None, department_id=None, printer_id=other_printer.id, date_from=None, date_to=None, db=db_session, actor=actor)
    assert printer_exc.value.status_code == 404


def test_job_listing_rejects_invalid_date_range(db_session: Session):
    actor = User(username="jobs-date-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(actor)
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        list_jobs(
            user_id=None,
            department_id=None,
            printer_id=None,
            date_from=datetime(2026, 6, 11, tzinfo=timezone.utc),
            date_to=datetime(2026, 6, 10, tzinfo=timezone.utc),
            db=db_session,
            actor=actor,
        )

    assert exc.value.status_code == 400


def test_same_user_and_printer_names_can_exist_in_different_organizations(db_session: Session):
    other_org = Organization(name="Cliente C", slug="cliente-c", is_active=True)
    db_session.add(other_org)
    db_session.flush()

    db_session.add_all(
        [
            User(
                organization_id=1,
                username="admin",
                full_name="Admin Default",
                password_hash=hash_password("AdminDefaultPassword2026"),
                role=UserRole.admin,
                is_active=True,
            ),
            User(
                organization_id=other_org.id,
                username="admin",
                full_name="Admin Cliente C",
                password_hash=hash_password("AdminClienteCPassword2026"),
                role=UserRole.admin,
                is_active=True,
            ),
            Printer(organization_id=1, name="KONICA", is_color=True),
            Printer(organization_id=other_org.id, name="KONICA", is_color=True),
        ]
    )
    db_session.commit()

    token = login(
        LoginRequest(username="admin", password="AdminClienteCPassword2026", organization_slug="cliente-c"),
        db=db_session,
    )

    assert db_session.query(User).filter(User.username == "admin").count() == 2
    assert db_session.query(Printer).filter(Printer.name == "KONICA").count() == 2
    assert token.role == "admin"
    assert token.organization_id == other_org.id
    assert token.organization_slug == "cliente-c"
    assert token.organization_name == "Cliente C"
    assert token.organization_billing_status == "trial"

    with pytest.raises(HTTPException) as exc:
        login(
            LoginRequest(username="admin", password="AdminClienteCPassword2026", organization_slug=None),
            db=db_session,
        )
    assert exc.value.status_code == 400


def test_login_normalizes_username_and_organization_slug(db_session: Session):
    organization = Organization(name="Cliente Login Normalizado", slug="cliente-login-normalizado", is_active=True)
    user = User(
        username="Admin-Login-Normalizado",
        full_name="Admin Login Normalizado",
        password_hash=hash_password("LoginNormalizadoPassword2026"),
        role=UserRole.admin,
        is_active=True,
        organization=organization,
    )
    db_session.add_all([organization, user])
    db_session.commit()

    token = login(
        LoginRequest(
            username=" ADMIN-LOGIN-NORMALIZADO ",
            password="LoginNormalizadoPassword2026",
            organization_slug=" CLIENTE-LOGIN-NORMALIZADO ",
        ),
        db=db_session,
    )

    assert token.role == "admin"
    assert token.organization_id == organization.id
    assert token.organization_slug == "cliente-login-normalizado"


def test_login_rejects_case_variant_duplicates_within_same_organization(db_session: Session):
    organization = Organization(name="Cliente Login Duplicado", slug="cliente-login-duplicado", is_active=True)
    db_session.add(organization)
    db_session.flush()
    db_session.add_all(
        [
            User(
                username="Admin.Legado",
                full_name="Admin Legado A",
                password_hash=hash_password("AdminLegadoPassword2026"),
                role=UserRole.admin,
                is_active=True,
                organization_id=organization.id,
            ),
            User(
                username="admin.legado",
                full_name="Admin Legado B",
                password_hash=hash_password("AdminLegadoPassword2026"),
                role=UserRole.admin,
                is_active=True,
                organization_id=organization.id,
            ),
        ]
    )
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        login(
            LoginRequest(
                username="ADMIN.LEGADO",
                password="AdminLegadoPassword2026",
                organization_slug="cliente-login-duplicado",
            ),
            db=db_session,
        )

    assert exc.value.status_code == 409
    assert "duplicado" in exc.value.detail.lower()


def test_organization_metrics_include_billable_monthly_jobs(db_session: Session):
    actor = User(username="org-metrics-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    user = User(username="org-metrics-user", full_name="Usuario", role=UserRole.user, is_active=True, organization_id=1)
    printer = Printer(organization_id=1, name="ORG METRICS PRINTER", is_color=True)
    db_session.add_all([actor, user, printer])
    db_session.flush()

    now = datetime.now(timezone.utc)
    previous_month = now - timedelta(days=40)
    db_session.add_all(
        [
            PrintJob(organization_id=1, user_id=user.id, printer_id=printer.id, pages=2, is_color=False, cost=0.10, status=JobStatus.authorized, submitted_at=now),
            PrintJob(organization_id=1, user_id=user.id, printer_id=printer.id, pages=3, is_color=True, cost=0.75, status=JobStatus.released, submitted_at=now),
            PrintJob(organization_id=1, user_id=user.id, printer_id=printer.id, pages=7, is_color=False, cost=0.35, status=JobStatus.pending_release, submitted_at=now),
            PrintJob(organization_id=1, user_id=user.id, printer_id=printer.id, pages=9, is_color=True, cost=2.25, status=JobStatus.blocked, submitted_at=now),
            PrintJob(organization_id=1, user_id=user.id, printer_id=printer.id, pages=5, is_color=False, cost=0.25, status=JobStatus.authorized, submitted_at=previous_month),
        ]
    )
    db_session.commit()

    organizations = list_organizations(db=db_session, actor=actor)
    default_org = next(organization for organization in organizations if organization.id == 1)

    assert default_org.jobs_count == 5
    assert default_org.jobs_month == 2
    assert default_org.pages_month == 5
    assert default_org.cost_month == 0.85
    assert default_org.pending_jobs_month == 1
    assert default_org.blocked_jobs_month == 1
    assert default_org.saved_pages_month == 9


def test_auth_context_returns_current_organization(db_session: Session):
    organization = Organization(name="Cliente Contexto", slug="cliente-contexto", is_active=True, billing_status="past_due")
    user = User(username="context-admin", full_name="Context Admin", role=UserRole.admin, is_active=True, organization=organization)
    db_session.add_all([organization, user])
    db_session.commit()

    context = current_auth_context(current_user=user)

    assert context.username == "context-admin"
    assert context.full_name == "Context Admin"
    assert context.role == "admin"
    assert context.organization_id == organization.id
    assert context.organization_slug == "cliente-contexto"
    assert context.organization_name == "Cliente Contexto"
    assert context.organization_billing_status == "past_due"


def test_agent_login_requires_explicit_organization_slug(db_session: Session):
    db_session.add(
        User(
            organization_id=1,
            username="agent-login-scope",
            full_name="Agent Login Scope",
            password_hash=hash_password("AgentLoginScopePassword2026"),
            role=UserRole.agent,
            is_active=True,
        )
    )
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        login(
            LoginRequest(username="agent-login-scope", password="AgentLoginScopePassword2026", organization_slug=None),
            db=db_session,
        )
    assert exc.value.status_code == 400

    token = login(
        LoginRequest(username="agent-login-scope", password="AgentLoginScopePassword2026", organization_slug="default"),
        db=db_session,
    )
    assert token.role == "agent"
    assert token.organization_id == 1


def test_inactive_organization_cannot_issue_login_token(db_session: Session):
    inactive_org = Organization(name="Cliente Inativo", slug="cliente-inativo", is_active=False)
    db_session.add(inactive_org)
    db_session.flush()
    inactive_user = User(
        organization_id=inactive_org.id,
        username="bloqueado",
        full_name="Bloqueado",
        password_hash=hash_password("BlockedOrgPassword2026"),
        role=UserRole.admin,
        is_active=True,
    )
    db_session.add(inactive_user)
    db_session.commit()

    with pytest.raises(HTTPException) as slug_exc:
        login(
            LoginRequest(username="bloqueado", password="BlockedOrgPassword2026", organization_slug="cliente-inativo"),
            db=db_session,
        )
    assert slug_exc.value.status_code == 401

    with pytest.raises(HTTPException) as no_slug_exc:
        login(
            LoginRequest(username="bloqueado", password="BlockedOrgPassword2026", organization_slug=None),
            db=db_session,
        )
    assert no_slug_exc.value.status_code == 401


def test_inactive_organization_rejects_existing_tokens_for_human_and_agent_users(db_session: Session):
    inactive_org = Organization(name="Cliente Token Inativo", slug="cliente-token-inativo", is_active=False)
    db_session.add(inactive_org)
    db_session.flush()
    admin = User(
        organization_id=inactive_org.id,
        username="admin-inativo-token",
        full_name="Admin Inativo",
        password_hash=hash_password("InactiveTokenAdminPassword2026"),
        role=UserRole.admin,
        is_active=True,
    )
    agent_user = User(
        organization_id=inactive_org.id,
        username="agent-inativo-token",
        full_name="Agent Inativo",
        password_hash=hash_password("InactiveTokenAgentPassword2026"),
        role=UserRole.agent,
        is_active=True,
    )
    db_session.add_all([admin, agent_user])
    db_session.commit()

    for user in [admin, agent_user]:
        token = create_access_token(user.username, {"role": user.role.value, "organization_id": user.organization_id})
        with pytest.raises(HTTPException) as token_exc:
            get_current_user(token, db_session)
        assert token_exc.value.status_code == 401


def test_suspended_organization_cannot_login_or_use_existing_token(db_session: Session):
    suspended_org = Organization(name="Cliente Suspenso", slug="cliente-suspenso", is_active=True, billing_status="suspended")
    db_session.add(suspended_org)
    db_session.flush()
    suspended_user = User(
        organization_id=suspended_org.id,
        username="suspenso",
        full_name="Suspenso",
        password_hash=hash_password("SuspendedOrgPassword2026"),
        role=UserRole.admin,
        is_active=True,
    )
    db_session.add(suspended_user)
    db_session.commit()

    with pytest.raises(HTTPException) as login_exc:
        login(
            LoginRequest(username="suspenso", password="SuspendedOrgPassword2026", organization_slug="cliente-suspenso"),
            db=db_session,
        )
    assert login_exc.value.status_code == 401

    token = create_access_token(
        suspended_user.username,
        {"role": suspended_user.role.value, "organization_id": suspended_user.organization_id},
    )
    with pytest.raises(HTTPException) as token_exc:
        get_current_user(token, db_session)
    assert token_exc.value.status_code == 401


def test_suspended_organization_rejects_agent_login_and_existing_token(db_session: Session):
    suspended_org = Organization(name="Cliente Agent Suspenso", slug="cliente-agent-suspenso", is_active=True, billing_status="suspended")
    db_session.add(suspended_org)
    db_session.flush()
    agent_user = User(
        organization_id=suspended_org.id,
        username="agent-suspenso",
        full_name="Agent Suspenso",
        password_hash=hash_password("SuspendedAgentPassword2026"),
        role=UserRole.agent,
        is_active=True,
    )
    db_session.add(agent_user)
    db_session.commit()

    with pytest.raises(HTTPException) as login_exc:
        login(
            LoginRequest(username="agent-suspenso", password="SuspendedAgentPassword2026", organization_slug="cliente-agent-suspenso"),
            db=db_session,
        )
    assert login_exc.value.status_code == 401

    token = create_access_token(agent_user.username, {"role": agent_user.role.value, "organization_id": agent_user.organization_id})
    with pytest.raises(HTTPException) as token_exc:
        get_current_user(token, db_session)
    assert token_exc.value.status_code == 401


def test_past_due_organization_can_still_login_until_suspended(db_session: Session):
    past_due_org = Organization(name="Cliente Em Atraso", slug="cliente-em-atraso", is_active=True, billing_status="past_due")
    db_session.add(past_due_org)
    db_session.flush()
    past_due_user = User(
        organization_id=past_due_org.id,
        username="atrasado",
        full_name="Atrasado",
        password_hash=hash_password("PastDueOrgPassword2026"),
        role=UserRole.admin,
        is_active=True,
    )
    db_session.add(past_due_user)
    db_session.commit()

    token = login(
        LoginRequest(username="atrasado", password="PastDueOrgPassword2026", organization_slug="cliente-em-atraso"),
        db=db_session,
    )

    assert token.organization_id == past_due_org.id
    assert get_current_user(token.access_token, db_session).id == past_due_user.id


def test_creating_organization_seeds_initial_admin_and_agent_users(db_session: Session):
    platform_admin = User(username="platform-admin", full_name="Platform Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(platform_admin)
    db_session.commit()

    created = create_organization(
        OrganizationCreate(
            name="Cliente Novo",
            slug="cliente-novo",
            billing_plan="professional",
            billing_status="active",
            contracted_printer_limit=25,
            admin_username="admin",
            admin_password="ClienteNovoAdminPassword2026",
            agent_username="agent",
            agent_password="ClienteNovoAgentPassword2026",
        ),
        db=db_session,
        actor=platform_admin,
    )

    seeded_users = db_session.query(User).filter(User.organization_id == created.id).order_by(User.username).all()
    assert [user.username for user in seeded_users] == ["admin", "agent"]
    assert {user.username: user.role for user in seeded_users} == {"admin": UserRole.admin, "agent": UserRole.agent}
    assert created.users_count == 2
    assert created.billing_plan == "professional"
    assert created.billing_status == "active"
    assert created.contracted_printer_limit == 25

    token = login(
        LoginRequest(username="admin", password="ClienteNovoAdminPassword2026", organization_slug="cliente-novo"),
        db=db_session,
    )
    assert token.organization_id == created.id

    audit = db_session.query(AuditLog).filter(AuditLog.action == "organization_created").one()
    assert audit.log_metadata["target_organization_id"] == created.id
    assert audit.log_metadata["admin_username"] == "admin"
    assert audit.log_metadata["agent_username"] == "agent"
    assert audit.log_metadata["billing_plan"] == "professional"
    assert audit.log_metadata["billing_status"] == "active"
    assert audit.log_metadata["contracted_printer_limit"] == 25
    assert "ClienteNovoAdminPassword2026" not in str(audit.log_metadata)
    assert "ClienteNovoAgentPassword2026" not in str(audit.log_metadata)


def test_creating_organization_normalizes_slug_and_initial_usernames(db_session: Session):
    platform_admin = User(username="platform-normalize-admin", full_name="Platform Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(platform_admin)
    db_session.commit()

    created = create_organization(
        OrganizationCreate(
            name="Cliente Normalizado",
            slug=" CLIENTE-NORMALIZADO ",
            admin_username=" Admin.Cliente ",
            admin_password="ClienteNormalizadoAdminPassword2026",
            agent_username=" Agent.Cliente ",
            agent_password="ClienteNormalizadoAgentPassword2026",
        ),
        db=db_session,
        actor=platform_admin,
    )

    seeded_users = db_session.query(User).filter(User.organization_id == created.id).order_by(User.username).all()
    audit = db_session.query(AuditLog).filter(AuditLog.action == "organization_created").one()
    token = login(
        LoginRequest(username=" admin.cliente ", password="ClienteNormalizadoAdminPassword2026", organization_slug=" CLIENTE-NORMALIZADO "),
        db=db_session,
    )

    assert created.slug == "cliente-normalizado"
    assert [user.username for user in seeded_users] == ["admin.cliente", "agent.cliente"]
    assert audit.log_metadata["slug"] == "cliente-normalizado"
    assert audit.log_metadata["admin_username"] == "admin.cliente"
    assert audit.log_metadata["agent_username"] == "agent.cliente"
    assert token.organization_id == created.id
    assert token.organization_slug == "cliente-normalizado"


def test_organization_create_rejects_default_or_shared_initial_passwords():
    with pytest.raises(ValidationError):
        OrganizationCreate(
            name="Cliente Senha Padrao",
            slug="cliente-senha-padrao",
            admin_password="AdminSeguro2026",
            agent_password="agent12345",
        )

    with pytest.raises(ValidationError):
        OrganizationCreate(
            name="Cliente Senha Igual",
            slug="cliente-senha-igual",
            admin_password="MesmaSenhaSegura2026",
            agent_password="MesmaSenhaSegura2026",
        )


def test_updating_organization_writes_changed_fields_to_audit(db_session: Session):
    platform_admin = User(username="platform-update-admin", full_name="Platform Admin", role=UserRole.admin, is_active=True, organization_id=1)
    organization = Organization(name="Cliente Audit Update", slug="cliente-audit-update", is_active=True, billing_plan="starter", billing_status="trial", contracted_printer_limit=0)
    db_session.add_all([platform_admin, organization])
    db_session.commit()

    updated = update_organization(
        organization.id,
        OrganizationUpdate(
            name="Cliente Audit Renomeado",
            is_active=False,
            billing_plan="enterprise",
            billing_status="past_due",
            contracted_printer_limit=75,
        ),
        db=db_session,
        actor=platform_admin,
    )
    assert updated.name == "Cliente Audit Renomeado"
    assert updated.is_active is False
    assert updated.billing_plan == "enterprise"
    assert updated.billing_status == "past_due"
    assert updated.contracted_printer_limit == 75

    audit = db_session.query(AuditLog).filter(AuditLog.action == "organization_updated").one()
    assert audit.log_metadata["target_organization_id"] == organization.id
    assert audit.log_metadata["target_organization_slug"] == "cliente-audit-update"
    assert audit.log_metadata["target_organization_name"] == "Cliente Audit Renomeado"
    assert audit.log_metadata["changes"]["name"] == {
        "before": "Cliente Audit Update",
        "after": "Cliente Audit Renomeado",
    }
    assert audit.log_metadata["changes"]["is_active"] == {"before": True, "after": False}
    assert audit.log_metadata["changes"]["billing_plan"] == {"before": "starter", "after": "enterprise"}
    assert audit.log_metadata["changes"]["billing_status"] == {"before": "trial", "after": "past_due"}
    assert audit.log_metadata["changes"]["contracted_printer_limit"] == {"before": 0, "after": 75}

    update_organization(
        organization.id,
        OrganizationUpdate(
            name="Cliente Audit Renomeado",
            is_active=False,
            billing_plan="enterprise",
            billing_status="past_due",
            contracted_printer_limit=75,
        ),
        db=db_session,
        actor=platform_admin,
    )
    assert db_session.query(AuditLog).filter(AuditLog.action == "organization_updated").count() == 1


def test_organization_printer_limit_cannot_be_reduced_below_active_printers(db_session: Session):
    platform_admin = User(username="platform-limit-admin", full_name="Platform Admin", role=UserRole.admin, is_active=True, organization_id=1)
    organization = Organization(name="Cliente Limite Reducao", slug="cliente-limite-reducao", is_active=True, contracted_printer_limit=5)
    db_session.add_all([platform_admin, organization])
    db_session.flush()
    db_session.add_all(
        [
            Printer(organization_id=organization.id, name="Limite Ativa A", is_color=False, is_active=True),
            Printer(organization_id=organization.id, name="Limite Ativa B", is_color=False, is_active=True),
            Printer(organization_id=organization.id, name="Limite Inativa", is_color=False, is_active=False),
        ]
    )
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        update_organization(
            organization.id,
            OrganizationUpdate(contracted_printer_limit=1),
            db=db_session,
            actor=platform_admin,
        )

    assert exc.value.status_code == 409
    assert "Limite contratado menor que as impressoras ativas atuais" in exc.value.detail
    db_session.rollback()
    unchanged = db_session.query(Organization).filter(Organization.id == organization.id).one()
    assert unchanged.contracted_printer_limit == 5
    assert db_session.query(AuditLog).filter(AuditLog.action == "organization_updated").count() == 0

    updated = update_organization(
        organization.id,
        OrganizationUpdate(contracted_printer_limit=2),
        db=db_session,
        actor=platform_admin,
    )
    assert updated.contracted_printer_limit == 2

    unlimited = update_organization(
        organization.id,
        OrganizationUpdate(contracted_printer_limit=0),
        db=db_session,
        actor=platform_admin,
    )
    assert unlimited.contracted_printer_limit == 0


def test_tenant_admin_cannot_update_own_organization_commercial_fields(db_session: Session):
    tenant_org = Organization(
        name="Cliente Read Only",
        slug="cliente-read-only",
        is_active=True,
        billing_plan="starter",
        billing_status="trial",
        contracted_printer_limit=10,
    )
    db_session.add(tenant_org)
    db_session.flush()
    tenant_admin = User(
        username="tenant-read-only-admin",
        full_name="Tenant Read Only Admin",
        role=UserRole.admin,
        is_active=True,
        organization_id=tenant_org.id,
    )
    db_session.add(tenant_admin)
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        update_organization(
            tenant_org.id,
            OrganizationUpdate(
                name="Cliente Alterado Pelo Tenant",
                billing_plan="enterprise",
                billing_status="active",
                contracted_printer_limit=100,
            ),
            db=db_session,
            actor=tenant_admin,
        )

    assert exc.value.status_code == 403
    db_session.rollback()
    unchanged = db_session.query(Organization).filter(Organization.id == tenant_org.id).one()
    assert unchanged.name == "Cliente Read Only"
    assert unchanged.billing_plan == "starter"
    assert unchanged.billing_status == "trial"
    assert unchanged.contracted_printer_limit == 10
    assert db_session.query(AuditLog).filter(AuditLog.action == "organization_updated").count() == 0


def test_admin_cannot_deactivate_own_organization(db_session: Session):
    default_org = db_session.query(Organization).filter(Organization.id == 1).one()
    default_org.is_active = True
    platform_admin = User(
        username="platform-lockout-admin",
        full_name="Platform Lockout Admin",
        role=UserRole.admin,
        is_active=True,
        organization_id=default_org.id,
    )
    db_session.add(platform_admin)
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        update_organization(
            default_org.id,
            OrganizationUpdate(is_active=False),
            db=db_session,
            actor=platform_admin,
        )
    assert exc.value.status_code == 400

    db_session.rollback()
    unchanged = db_session.query(Organization).filter(Organization.id == default_org.id).one()
    assert unchanged.is_active is True
    assert db_session.query(AuditLog).filter(AuditLog.action == "organization_updated").count() == 0


def test_admin_cannot_suspend_own_organization(db_session: Session):
    default_org = db_session.query(Organization).filter(Organization.id == 1).one()
    default_org.billing_status = "active"
    platform_admin = User(
        username="platform-self-suspend-admin",
        full_name="Platform Self Suspend Admin",
        role=UserRole.admin,
        is_active=True,
        organization_id=default_org.id,
    )
    db_session.add(platform_admin)
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        update_organization(
            default_org.id,
            OrganizationUpdate(billing_status="suspended"),
            db=db_session,
            actor=platform_admin,
        )
    assert exc.value.status_code == 400

    db_session.rollback()
    unchanged = db_session.query(Organization).filter(Organization.id == default_org.id).one()
    assert unchanged.billing_status == "active"
    assert db_session.query(AuditLog).filter(AuditLog.action == "organization_updated").count() == 0


def test_organization_list_includes_scoped_usage_counts(db_session: Session):
    default_org = db_session.query(Organization).filter(Organization.id == 1).one()
    default_org.contracted_printer_limit = 2
    other_org = Organization(name="Cliente Usage", slug="cliente-usage", is_active=True, contracted_printer_limit=1)
    admin = User(username="usage-admin", full_name="Usage Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add_all([other_org, admin])
    db_session.flush()

    default_user = User(username="usage-user-a", full_name="Usage A", role=UserRole.user, is_active=True, organization_id=1)
    default_printer = Printer(name="Usage Printer A", is_color=False, organization_id=1)
    default_inactive_printer = Printer(name="Usage Printer Inativa A", is_color=False, is_active=False, organization_id=1)
    default_agent = PrintAgent(
        agent_uid="usage-agent-a",
        computer_name="PC-A",
        organization_id=1,
        last_seen_at=datetime.now(timezone.utc),
    )
    other_user = User(username="usage-user-b", full_name="Usage B", role=UserRole.user, is_active=True, organization_id=other_org.id)
    other_printer = Printer(name="Usage Printer B", is_color=False, organization_id=other_org.id)
    other_agent = PrintAgent(
        agent_uid="usage-agent-b",
        computer_name="PC-B",
        organization_id=other_org.id,
        last_seen_at=datetime.now(timezone.utc) - timedelta(minutes=10),
    )
    db_session.add_all([default_user, default_printer, default_inactive_printer, default_agent, other_user, other_printer, other_agent])
    db_session.flush()
    db_session.add_all(
        [
            PrintJob(
                organization_id=1,
                user_id=default_user.id,
                printer_id=default_printer.id,
                agent_id=default_agent.id,
                pages=1,
                is_color=False,
                cost=0.05,
                status=JobStatus.authorized,
                submitted_at=datetime.now(timezone.utc),
            ),
            PrintJob(
                organization_id=1,
                user_id=default_user.id,
                printer_id=default_printer.id,
                agent_id=default_agent.id,
                pages=99,
                is_color=False,
                cost=9.90,
                status=JobStatus.blocked,
                submitted_at=datetime.now(timezone.utc),
            ),
            PrintJob(
                organization_id=1,
                user_id=other_user.id,
                printer_id=other_printer.id,
                agent_id=default_agent.id,
                pages=50,
                is_color=False,
                cost=5.00,
                status=JobStatus.authorized,
                submitted_at=datetime.now(timezone.utc),
            ),
            PrintJob(
                organization_id=other_org.id,
                user_id=other_user.id,
                printer_id=other_printer.id,
                agent_id=other_agent.id,
                pages=2,
                is_color=False,
                cost=0.10,
                status=JobStatus.authorized,
                submitted_at=datetime.now(timezone.utc),
            ),
        ]
    )
    db_session.commit()

    rows = list_organizations(db=db_session, actor=admin)
    default_row = next(row for row in rows if row.id == 1)
    other_row = next(row for row in rows if row.id == other_org.id)

    assert default_row.users_count >= 2
    assert default_row.printers_count >= 2
    assert default_row.active_printers_count >= 1
    assert default_row.contracted_printer_usage_percent >= 50.0
    assert default_row.contracted_printer_limit_status in {"ok", "warning"}
    assert default_row.agents_count >= 1
    assert default_row.online_agents_count >= 1
    assert default_row.jobs_count == 2
    assert default_row.pages_month == 1
    assert default_row.cost_month == 0.05
    assert default_row.pending_jobs_month == 0
    assert default_row.blocked_jobs_month == 1
    assert default_row.saved_pages_month == 99
    assert other_row.users_count == 1
    assert other_row.printers_count == 1
    assert other_row.active_printers_count == 1
    assert other_row.contracted_printer_usage_percent == 100.0
    assert other_row.contracted_printer_limit_status == "warning"
    assert other_row.agents_count == 1
    assert other_row.online_agents_count == 0
    assert other_row.offline_agents_count == 1
    assert other_row.jobs_count == 1
    assert other_row.pages_month == 2
    assert other_row.cost_month == 0.10
    assert other_row.pending_jobs_month == 0
    assert other_row.blocked_jobs_month == 0
    assert other_row.saved_pages_month == 0


def test_organization_list_marks_printer_contract_overage(db_session: Session):
    organization = Organization(name="Cliente Overage", slug="cliente-overage", is_active=True, contracted_printer_limit=1)
    admin = User(username="overage-admin", full_name="Overage Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add_all([organization, admin])
    db_session.flush()
    db_session.add_all(
        [
            Printer(name="Overage Printer A", is_color=False, organization_id=organization.id),
            Printer(name="Overage Printer B", is_color=False, organization_id=organization.id),
        ]
    )
    db_session.commit()

    rows = list_organizations(db=db_session, actor=admin)
    row = next(item for item in rows if item.id == organization.id)

    assert row.active_printers_count == 2
    assert row.contracted_printer_usage_percent == 200.0
    assert row.contracted_printer_limit_status == "exceeded"


def test_department_admin_crud_is_scoped_and_protects_in_use_departments(db_session: Session):
    other_org = Organization(name="Cliente Dept", slug="cliente-dept", is_active=True)
    admin = User(username="dept-admin", full_name="Dept Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add_all([other_org, admin])
    db_session.flush()

    other_department = Department(organization_id=other_org.id, name="Financeiro")
    db_session.add(other_department)
    db_session.commit()

    department = create_department(DepartmentCreate(name="Financeiro", cost_center="CC-FIN"), db=db_session, actor=admin)
    assert department.organization_id == 1
    assert department.cost_center == "CC-FIN"
    assert [item.name for item in list_departments(db=db_session, actor=admin)] == ["Financeiro"]

    updated = update_department(department.id, DepartmentUpdate(name="Administrativo", cost_center="CC-ADM"), db=db_session, actor=admin)
    assert updated.name == "Administrativo"
    assert updated.cost_center == "CC-ADM"
    assert db_session.query(Department).filter(Department.organization_id == other_org.id, Department.name == "Financeiro").one()

    user = User(username="dept-user", full_name="Dept User", role=UserRole.user, is_active=True, organization_id=1, department_id=department.id)
    db_session.add(user)
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        delete_department(department.id, db=db_session, actor=admin)
    assert exc.value.status_code == 400

    user.department_id = None
    db_session.commit()
    result = delete_department(department.id, db=db_session, actor=admin)
    assert result == {"status": "deleted"}


def test_department_update_and_delete_audit_metadata(db_session: Session):
    admin = User(username="dept-audit-admin", full_name="Dept Audit Admin", role=UserRole.admin, is_active=True, organization_id=1)
    department = Department(organization_id=1, name="Financeiro", cost_center="CC-FIN")
    db_session.add_all([admin, department])
    db_session.commit()

    updated = update_department(department.id, DepartmentUpdate(name="Operacoes", cost_center="CC-OPS"), db=db_session, actor=admin)
    assert updated.name == "Operacoes"
    assert updated.cost_center == "CC-OPS"
    update_audit = db_session.query(AuditLog).filter(AuditLog.action == "department_updated", AuditLog.entity_id == department.id).one()
    assert update_audit.log_metadata["changes"] == {
        "name": {"before": "Financeiro", "after": "Operacoes"},
        "cost_center": {"before": "CC-FIN", "after": "CC-OPS"},
    }

    result = delete_department(department.id, db=db_session, actor=admin)
    assert result == {"status": "deleted"}
    delete_audit = db_session.query(AuditLog).filter(AuditLog.action == "department_deleted", AuditLog.entity_id == department.id).one()
    assert delete_audit.log_metadata == {"name": "Operacoes", "cost_center": "CC-OPS"}


def test_user_department_id_assignment_is_scoped_and_clearable(db_session: Session):
    other_org = Organization(name="Cliente User Dept", slug="cliente-user-dept", is_active=True)
    admin = User(username="user-dept-admin", full_name="User Dept Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add_all([other_org, admin])
    db_session.flush()

    department = Department(organization_id=1, name="Financeiro")
    other_department = Department(organization_id=other_org.id, name="Financeiro")
    db_session.add_all([department, other_department])
    db_session.commit()

    created = create_user_endpoint(
        UserCreate(
            username="com-depto",
            full_name="Com Departamento",
            department_id=department.id,
            monthly_limit=500,
            monthly_balance=50.0,
        ),
        db=db_session,
        actor=admin,
    )
    assert created.department_id == department.id
    assert created.department_name == "Financeiro"

    with pytest.raises(HTTPException) as exc:
        create_user_endpoint(
            UserCreate(username="outro-depto", full_name="Outro Departamento", department_id=other_department.id),
            db=db_session,
            actor=admin,
        )
    assert exc.value.status_code == 404

    cleared = update_user_endpoint(created.id, UserUpdate(department_id=None), db=db_session, actor=admin)
    assert cleared.department_id is None
    assert cleared.department_name is None


def test_user_creation_normalizes_username_and_rejects_case_duplicate(db_session: Session):
    admin = User(username="user-normalize-admin", full_name="User Normalize Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(admin)
    db_session.commit()

    created = create_user_endpoint(
        UserCreate(username=" DIEGO.LCD ", full_name="Diego LCD"),
        db=db_session,
        actor=admin,
    )

    assert created.username == "diego.lcd"
    assert db_session.query(User).filter(User.organization_id == admin.organization_id, User.username == "diego.lcd").count() == 1

    with pytest.raises(HTTPException) as exc:
        create_user_endpoint(
            UserCreate(username="diego.LCD", full_name="Diego Duplicado"),
            db=db_session,
            actor=admin,
        )

    assert exc.value.status_code == 409
    db_session.rollback()
    assert db_session.query(User).filter(User.organization_id == admin.organization_id, func.lower(User.username) == "diego.lcd").count() == 1


def test_user_update_audit_includes_changed_fields(db_session: Session):
    admin = User(username="audit-user-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    department = Department(organization_id=1, name="Financeiro")
    new_department = Department(organization_id=1, name="Operacoes")
    user = User(username="audit-user", full_name="Usuario Antigo", role=UserRole.user, is_active=True, organization_id=1, department=department)
    db_session.add_all([admin, department, new_department, user])
    db_session.flush()
    quota = Quota(
        organization_id=1,
        user_id=user.id,
        year=datetime.now(timezone.utc).year,
        month=datetime.now(timezone.utc).month,
        monthly_limit=500,
        monthly_balance=50.0,
        used_pages=0,
        used_balance=0.0,
    )
    db_session.add(quota)
    db_session.commit()

    update_user_endpoint(
        user.id,
        UserUpdate(full_name="Usuario Novo", department_id=new_department.id, is_active=False, monthly_limit=750, monthly_balance=75.0),
        db=db_session,
        actor=admin,
    )

    audit = db_session.query(AuditLog).filter(AuditLog.action == "user_updated", AuditLog.entity_id == user.id).one()
    assert audit.log_metadata["changes"]["full_name"] == {"before": "Usuario Antigo", "after": "Usuario Novo"}
    assert audit.log_metadata["changes"]["department_id"] == {"before": department.id, "after": new_department.id}
    assert audit.log_metadata["changes"]["department_name"] == {"before": "Financeiro", "after": "Operacoes"}
    assert audit.log_metadata["changes"]["is_active"] == {"before": True, "after": False}
    assert audit.log_metadata["changes"]["monthly_limit"] == {"before": 500, "after": 750}
    assert audit.log_metadata["changes"]["monthly_balance"] == {"before": 50.0, "after": 75.0}


def test_quota_update_audit_includes_limit_change(db_session: Session):
    manager = User(username="quota-manager", full_name="Manager", role=UserRole.manager, is_active=True, organization_id=1)
    user = User(username="quota-user", full_name="Usuario", role=UserRole.user, is_active=True, organization_id=1)
    db_session.add_all([manager, user])
    db_session.flush()
    quota = Quota(
        organization_id=1,
        user_id=user.id,
        year=2026,
        month=6,
        monthly_limit=500,
        monthly_balance=50.0,
        used_pages=120,
        used_balance=12.0,
    )
    db_session.add(quota)
    db_session.commit()

    updated = update_quota(quota.id, QuotaUpdate(monthly_limit=750), db=db_session, actor=manager)

    assert updated.monthly_limit == 750
    assert updated.remaining_pages == 630
    audit = db_session.query(AuditLog).filter(AuditLog.action == "quota_updated", AuditLog.entity_id == quota.id).one()
    assert audit.actor_user_id == manager.id
    assert audit.log_metadata["username"] == "quota-user"
    assert audit.log_metadata["period"] == "2026-06"
    assert audit.log_metadata["changes"]["monthly_limit"] == {"before": 500, "after": 750}


def test_printer_update_audit_includes_changed_fields(db_session: Session):
    admin = User(username="audit-printer-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    printer = Printer(
        organization_id=1,
        name="KONICA ANTIGA",
        location="Recepcao",
        ip_address="192.168.1.10",
        is_color=True,
        cost_mono=0.05,
        cost_color=0.25,
        is_active=True,
    )
    db_session.add_all([admin, printer])
    db_session.commit()

    update_printer_endpoint(
        printer.id,
        PrinterUpdate(name="KONICA NOVA", ip_address="", cost_color=0.30, is_active=False),
        db=db_session,
        actor=admin,
    )

    audit = db_session.query(AuditLog).filter(AuditLog.action == "printer_updated", AuditLog.entity_id == printer.id).one()
    assert audit.log_metadata["changes"]["name"] == {"before": "KONICA ANTIGA", "after": "KONICA NOVA"}
    assert audit.log_metadata["changes"]["ip_address"] == {"before": "192.168.1.10", "after": None}
    assert audit.log_metadata["changes"]["cost_color"] == {"before": 0.25, "after": 0.30}
    assert audit.log_metadata["changes"]["is_active"] == {"before": True, "after": False}
    assert "cost_mono" not in audit.log_metadata["changes"]


def test_admin_can_manage_human_user_roles(db_session: Session):
    admin = User(username="role-admin", full_name="Role Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(admin)
    db_session.commit()

    created = create_user_endpoint(
        UserCreate(
            username="gestor-financeiro",
            full_name="Gestor Financeiro",
            role=UserRole.manager,
            password="GestorFinanceiroPassword2026",
        ),
        db=db_session,
        actor=admin,
    )
    assert created.role == UserRole.manager

    updated = update_user_endpoint(created.id, UserUpdate(role=UserRole.admin), db=db_session, actor=admin)
    assert updated.role == UserRole.admin


def test_panel_user_creation_requires_password(db_session: Session):
    admin = User(username="panel-password-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(admin)
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        create_user_endpoint(
            UserCreate(username="gestor-sem-senha", full_name="Gestor Sem Senha", role=UserRole.manager),
            db=db_session,
            actor=admin,
        )

    assert exc.value.status_code == 422
    assert db_session.query(User).filter(User.username == "gestor-sem-senha").first() is None


def test_panel_user_creation_rejects_unsafe_password(db_session: Session):
    admin = User(username="panel-unsafe-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(admin)
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        create_user_endpoint(
            UserCreate(
                username="gestor-senha-fraca",
                full_name="Gestor Senha Fraca",
                role=UserRole.manager,
                password="admin12345",
            ),
            db=db_session,
            actor=admin,
        )

    assert exc.value.status_code == 400
    assert db_session.query(User).filter(User.username == "gestor-senha-fraca").first() is None


def test_panel_user_creation_hashes_password(db_session: Session):
    admin = User(username="panel-hash-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(admin)
    db_session.commit()

    created = create_user_endpoint(
        UserCreate(
            username="gestor-com-senha",
            full_name="Gestor Com Senha",
            role=UserRole.manager,
            password="GestorComSenhaPassword2026",
        ),
        db=db_session,
        actor=admin,
    )

    saved_user = db_session.get(User, created.id)
    assert saved_user is not None
    assert saved_user.password_hash is not None
    assert verify_password("GestorComSenhaPassword2026", saved_user.password_hash)


def test_promoting_user_to_panel_role_requires_password_when_missing(db_session: Session):
    admin = User(username="promote-panel-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    user = User(username="promote-panel-user", full_name="Promover Usuario", role=UserRole.user, is_active=True, organization_id=1)
    db_session.add_all([admin, user])
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        update_user_endpoint(user.id, UserUpdate(role=UserRole.manager), db=db_session, actor=admin)

    assert exc.value.status_code == 422
    db_session.refresh(user)
    assert user.role == UserRole.user
    assert user.password_hash is None

    updated = update_user_endpoint(
        user.id,
        UserUpdate(role=UserRole.manager, password="PromotedUserPassword2026"),
        db=db_session,
        actor=admin,
    )

    assert updated.role == UserRole.manager
    db_session.refresh(user)
    assert verify_password("PromotedUserPassword2026", user.password_hash)


def test_user_endpoint_rejects_agent_role_creation(db_session: Session):
    admin = User(username="agent-create-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(admin)
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        create_user_endpoint(
            UserCreate(username="agent-manual", full_name="Agent Manual", role=UserRole.agent),
            db=db_session,
            actor=admin,
        )

    assert exc.value.status_code == 400
    assert db_session.query(User).filter(User.username == "agent-manual").first() is None


def test_user_endpoint_rejects_agent_role_changes(db_session: Session):
    admin = User(username="agent-role-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    user = User(username="humano", full_name="Humano", role=UserRole.user, is_active=True, organization_id=1)
    agent_user = User(username="coletor", full_name="Coletor", role=UserRole.agent, is_active=True, organization_id=1)
    db_session.add_all([admin, user, agent_user])
    db_session.commit()

    with pytest.raises(HTTPException) as demote_exc:
        update_user_endpoint(agent_user.id, UserUpdate(role=UserRole.user), db=db_session, actor=admin)
    assert demote_exc.value.status_code == 400

    with pytest.raises(HTTPException) as promote_exc:
        update_user_endpoint(user.id, UserUpdate(role=UserRole.agent), db=db_session, actor=admin)
    assert promote_exc.value.status_code == 400

    db_session.refresh(user)
    db_session.refresh(agent_user)
    assert user.role == UserRole.user
    assert agent_user.role == UserRole.agent


def test_seed_agent_user_cannot_be_deleted_by_accident(db_session: Session):
    admin = User(username="delete-agent-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    seed_agent = User(username="agent", full_name="Agent Padrao", role=UserRole.agent, is_active=True, organization_id=1)
    db_session.add_all([admin, seed_agent])
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        delete_user_endpoint(seed_agent.id, db=db_session, actor=admin)

    assert exc.value.status_code == 400
    assert db_session.get(User, seed_agent.id) is not None


def test_enrolled_agent_user_can_be_deleted(db_session: Session):
    admin = User(username="delete-enrolled-agent-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    custom_agent = User(username="agent-desktop-5jvqbpu-38f611", full_name="Agent Filial", role=UserRole.agent, is_active=True, organization_id=1)
    db_session.add_all([admin, custom_agent])
    db_session.commit()

    result = delete_user_endpoint(custom_agent.id, db=db_session, actor=admin)

    assert result["status"] == "deleted"
    assert db_session.get(User, custom_agent.id) is None


def test_user_with_linked_policy_cannot_be_deleted(db_session: Session):
    admin = User(username="delete-policy-user-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    user = User(username="delete-policy-user", full_name="Usuario com Politica", role=UserRole.user, is_active=True, organization_id=1)
    db_session.add_all([admin, user])
    db_session.flush()
    policy = PrintPolicy(
        organization_id=1,
        name="Excecao usuario especifico",
        priority=10,
        rule_type=PolicyRuleType.always,
        action=PolicyAction.allow,
        user_id=user.id,
    )
    db_session.add(policy)
    db_session.commit()

    with pytest.raises(HTTPException) as exc:
        delete_user_endpoint(user.id, db=db_session, actor=admin)

    assert exc.value.status_code == 409
    assert "politicas vinculadas" in exc.value.detail
    assert db_session.get(User, user.id) is not None
    assert db_session.get(PrintPolicy, policy.id) is not None


def test_delete_user_detaches_requested_queue_actions(db_session: Session):
    admin = User(username="delete-queue-user-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    user = User(username="delete-queue-user", full_name="Usuario com Acao", role=UserRole.user, is_active=True, organization_id=1)
    printer = Printer(organization_id=1, name="KONICA USER ACTION", is_color=True)
    agent = PrintAgent(organization_id=1, agent_uid="agent-delete-queue-user", computer_name="PC-QUEUE-USER")
    db_session.add_all([admin, user, printer, agent])
    db_session.flush()
    action = AgentQueueAction(
        organization_id=1,
        agent_id=agent.id,
        printer_id=printer.id,
        requested_by_user_id=user.id,
        action_type=AgentQueueActionType.create_queue,
        queue_name="KONICA Usuario",
        driver_name="KONICA Driver",
        ip_address="192.168.1.125",
        status=AgentQueueActionStatus.pending,
    )
    db_session.add(action)
    db_session.commit()

    result = delete_user_endpoint(user.id, db=db_session, actor=admin)
    db_session.refresh(action)
    audit = db_session.query(AuditLog).filter(AuditLog.action == "user_deleted", AuditLog.entity_id == user.id).one()

    assert result == {"status": "deleted", "deleted_jobs": 0, "detached_queue_actions": 1}
    assert action.requested_by_user_id is None
    assert db_session.get(User, user.id) is None
    assert audit.log_metadata["detached_queue_actions"] == 1


def test_agent_user_password_can_be_rotated_without_role_change(db_session: Session):
    admin = User(username="rotate-agent-admin", full_name="Admin", role=UserRole.admin, is_active=True, organization_id=1)
    agent_user = User(
        username="rotate-agent",
        full_name="Agent Rotacionado",
        role=UserRole.agent,
        is_active=True,
        organization_id=1,
        password_hash=hash_password("AgentOldPassword2026"),
    )
    db_session.add_all([admin, agent_user])
    db_session.commit()

    updated = update_user_endpoint(
        agent_user.id,
        UserUpdate(full_name="Agent Rotacionado", password="AgentNewPassword2026"),
        db=db_session,
        actor=admin,
    )

    assert updated.role == UserRole.agent
    db_session.refresh(agent_user)
    assert verify_password("AgentNewPassword2026", agent_user.password_hash)
    audit = db_session.query(AuditLog).filter(AuditLog.action == "user_updated", AuditLog.entity_id == agent_user.id).one()
    assert audit.log_metadata["changes"]["password"] == {"before": False, "after": True}

    with pytest.raises(HTTPException) as exc:
        update_user_endpoint(agent_user.id, UserUpdate(password="agent12345"), db=db_session, actor=admin)
    assert exc.value.status_code == 400
    db_session.refresh(agent_user)
    assert verify_password("AgentNewPassword2026", agent_user.password_hash)


def test_audit_log_listing_is_scoped_by_organization(db_session: Session):
    other_org = Organization(name="Cliente Audit", slug="cliente-audit", is_active=True)
    admin = User(username="audit-admin", full_name="Audit Admin", role=UserRole.admin, is_active=True, organization_id=1)
    other_admin = User(username="audit-other", full_name="Audit Other", role=UserRole.admin, is_active=True, organization=other_org)
    db_session.add_all([other_org, admin, other_admin])
    db_session.flush()

    write_audit(
        db_session,
        action="printer_updated",
        entity="printers",
        entity_id=10,
        actor_user_id=admin.id,
        metadata={"name": "KONICA"},
    )
    write_audit(
        db_session,
        action="printer_updated",
        entity="printers",
        entity_id=20,
        actor_user_id=other_admin.id,
        metadata={"name": "OUTRA"},
    )
    db_session.commit()

    logs = list_audit_logs(action=None, entity=None, date_from=None, date_to=None, limit=100, db=db_session, actor=admin)
    assert [log.entity_id for log in logs] == [10]
    assert logs[0].actor_username == "audit-admin"
    assert logs[0].metadata == {"name": "KONICA"}

    filtered = list_audit_logs(
        action="printer_updated",
        entity="printers",
        date_from=None,
        date_to=None,
        limit=100,
        db=db_session,
        actor=admin,
    )
    assert [log.entity_id for log in filtered] == [10]


def test_audit_log_actor_name_does_not_cross_organizations(db_session: Session):
    other_org = Organization(name="Cliente Audit Ator", slug="cliente-audit-ator", is_active=True)
    admin = User(username="audit-main-admin", full_name="Audit Main", role=UserRole.admin, is_active=True, organization_id=1)
    other_admin = User(username="audit-leaked-admin", full_name="Audit Other", role=UserRole.admin, is_active=True, organization=other_org)
    db_session.add_all([other_org, admin, other_admin])
    db_session.flush()

    write_audit(
        db_session,
        action="maintenance_imported",
        entity="audit_logs",
        entity_id=123,
        actor_user_id=other_admin.id,
        organization_id=admin.organization_id,
        metadata={"source": "manual"},
    )
    db_session.commit()

    logs = list_audit_logs(action=None, entity=None, date_from=None, date_to=None, limit=100, db=db_session, actor=admin)

    assert [log.entity_id for log in logs] == [123]
    assert logs[0].actor_user_id == other_admin.id
    assert logs[0].actor_username is None


def test_audit_log_facets_are_scoped_by_organization(db_session: Session):
    other_org = Organization(name="Cliente Facets", slug="cliente-facets", is_active=True)
    admin = User(username="audit-facets", full_name="Audit Facets", role=UserRole.admin, is_active=True, organization_id=1)
    other_admin = User(username="audit-facets-other", full_name="Audit Other", role=UserRole.admin, is_active=True, organization=other_org)
    db_session.add_all([other_org, admin, other_admin])
    db_session.flush()

    write_audit(
        db_session,
        action="settings_updated",
        entity="settings",
        actor_user_id=admin.id,
    )
    write_audit(
        db_session,
        action="printer_created",
        entity="printers",
        actor_user_id=admin.id,
    )
    write_audit(
        db_session,
        action="organization_updated",
        entity="organizations",
        actor_user_id=other_admin.id,
    )
    db_session.commit()

    facets = list_audit_log_facets(db=db_session, actor=admin)

    assert facets.actions == ["printer_created", "settings_updated"]
    assert facets.entities == ["printers", "settings"]


def test_audit_log_filters_by_date_and_exports_csv(db_session: Session):
    admin = User(username="=audit-export", full_name="Audit Export", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(admin)
    db_session.flush()

    old_log = write_audit(
        db_session,
        action="user_created",
        entity="users",
        entity_id=1,
        actor_user_id=admin.id,
        metadata={"username": "antigo"},
    )
    old_log.created_at = datetime(2026, 1, 1, 10, 0, tzinfo=timezone.utc)
    new_log = write_audit(
        db_session,
        action="printer_updated",
        entity="printers",
        entity_id=2,
        actor_user_id=admin.id,
        metadata={"name": "KONICA"},
    )
    new_log.created_at = datetime(2026, 2, 1, 10, 0, tzinfo=timezone.utc)
    db_session.commit()

    filtered = list_audit_logs(
        action=None,
        entity=None,
        date_from=datetime(2026, 2, 1, 0, 0, tzinfo=timezone.utc),
        date_to=datetime(2026, 2, 28, 23, 59, tzinfo=timezone.utc),
        limit=100,
        db=db_session,
        actor=admin,
    )
    assert [log.entity_id for log in filtered] == [2]

    response = export_audit_logs(
        action=None,
        entity=None,
        date_from=datetime(2026, 2, 1, 0, 0, tzinfo=timezone.utc),
        date_to=datetime(2026, 2, 28, 23, 59, tzinfo=timezone.utc),
        limit=100,
        db=db_session,
        actor=admin,
    )
    body = response.body.decode("utf-8")
    assert "data_hora,ator,acao,entidade,id_entidade,detalhes" in body
    assert "'=audit-export,printer_updated,printers,2" in body
    assert "\n=audit-export" not in body
    assert "KONICA" in body
    assert "antigo" not in body
    assert "audit_logs_exported" not in body
    assert response.headers["content-disposition"] == 'attachment; filename="auditoria-2026-02-01-2026-02-28.csv"'

    audit = db_session.query(AuditLog).filter(AuditLog.action == "audit_logs_exported").one()
    assert audit.actor_user_id == admin.id
    assert audit.log_metadata["filename"] == "auditoria-2026-02-01-2026-02-28.csv"
    assert audit.log_metadata["rows"] == 1
    assert audit.log_metadata["total_matching_rows"] == 1
    assert audit.log_metadata["limit"] == 100
    assert audit.log_metadata["truncated"] is False
    assert audit.log_metadata["filters"]["date_from"] == "2026-02-01T00:00:00+00:00"
    assert audit.log_metadata["filters"]["date_to"] == "2026-02-28T23:59:00+00:00"


def test_audit_log_export_audit_marks_truncated_result(db_session: Session):
    admin = User(username="audit-truncated-admin", full_name="Audit Truncated", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(admin)
    db_session.flush()

    for index in range(3):
        write_audit(
            db_session,
            action="printer_updated",
            entity="printers",
            entity_id=index + 1,
            actor_user_id=admin.id,
            metadata={"index": index},
        )
    db_session.commit()

    response = export_audit_logs(
        action="printer_updated",
        entity=None,
        date_from=None,
        date_to=None,
        limit=1,
        db=db_session,
        actor=admin,
    )

    body_rows = [line for line in response.body.decode("utf-8").splitlines() if line]
    assert len(body_rows) == 2
    audit = db_session.query(AuditLog).filter(AuditLog.action == "audit_logs_exported").one()
    assert audit.log_metadata["filename"] == "auditoria.csv"
    assert audit.log_metadata["rows"] == 1
    assert audit.log_metadata["total_matching_rows"] == 3
    assert audit.log_metadata["limit"] == 1
    assert audit.log_metadata["truncated"] is True
    assert audit.log_metadata["filters"]["action"] == "printer_updated"


def test_audit_log_rejects_invalid_date_range_without_export_audit(db_session: Session):
    admin = User(username="audit-date-admin", full_name="Audit Date", role=UserRole.admin, is_active=True, organization_id=1)
    db_session.add(admin)
    db_session.commit()
    date_from = datetime(2026, 3, 1, tzinfo=timezone.utc)
    date_to = datetime(2026, 2, 1, tzinfo=timezone.utc)

    with pytest.raises(HTTPException) as list_exc:
        list_audit_logs(action=None, entity=None, date_from=date_from, date_to=date_to, limit=100, db=db_session, actor=admin)
    assert list_exc.value.status_code == 400

    with pytest.raises(HTTPException) as export_exc:
        export_audit_logs(action=None, entity=None, date_from=date_from, date_to=date_to, limit=100, db=db_session, actor=admin)
    assert export_exc.value.status_code == 400
    assert db_session.query(AuditLog).filter(AuditLog.action == "audit_logs_exported").count() == 0
